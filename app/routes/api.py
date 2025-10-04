"""
API endpoints - wszystkie endpointy /api/*
Zachowuje pełną kompatybilność z istniejącymi wywołaniami frontend
"""

import json
import logging
import os
from flask import Blueprint, request, jsonify, session
from ..auth import login_required, admin_required, rate_limit_required
from flask_limiter import Limiter
from flask_limiter.util import get_remote_address
from ..database import get_db
from ..cache import cache, cached, invalidate_cache_pattern
from ..performance import monitor_performance, perf_counter

logger = logging.getLogger(__name__)

bp = Blueprint('api', __name__, url_prefix='/api')

# Rate limiter dla API - optimized limits
limiter = Limiter(
    key_func=get_remote_address,
    default_limits=["200 per minute"]  # Increased from 100 to 200 for better UX
)

def safe_get_json():
    """Bezpieczne pobieranie JSON z requestu"""
    try:
        return request.get_json() or {}
    except Exception as e:
        logger.error(f"Błąd podczas parsowania JSON: {e}")
        return {}

# ============================================================================
# ENDPOINTY ZARZĄDZANIA ZMIANAMI
# ============================================================================

@bp.post("/save")
@login_required
@limiter.limit("20 per minute")  # Increased from 10 to 20 for better UX
def api_save():
    """Zapisuje zmiany w grafiku"""
    try:
        from ..database import ensure_schedule_changes_table
        ensure_schedule_changes_table()  # Upewnij się, że tabela istnieje
        logger.info(f"Przetwarzanie {len(request.get_json() or {})} zmian w api_save")
        
        # Pobierz nazwę użytkownika
        user_email = session.get('user_email', 'nieznany')
        user_role = session.get('user_role', 'USER')
        
        # Pobierz dane z requestu
        data = safe_get_json()
        changes = data.get('changes', [])
        is_draft = data.get('is_draft', False)  # Nowy parametr dla trybu draft
        
        if not changes:
            return jsonify(error="Brak zmian do zapisania"), 400
        
        # Sprawdź czy użytkownik może używać trybu draft
        if is_draft and user_role != 'ADMIN':
            return jsonify(error="Tylko administratorzy mogą używać trybu roboczego"), 403
        
        db = get_db()
        
        # Jeśli to tryb draft, użyj specjalnej logiki
        if is_draft:
            from ..database import ensure_draft_shifts_table
            ensure_draft_shifts_table()
            
            user_id = session.get('user_id')
            
            # Wyczyść wszystkie istniejące drafty dla tego użytkownika
            db.execute("DELETE FROM draft_shifts WHERE created_by = ?", (user_id,))
            
            # Zapisz nowe zmiany w trybie draft
            for change in changes:
                date = change.get('date')
                employee = change.get('employee') or change.get('name')
                shift_type_raw = change.get('shift_type') or change.get('value', '')
                
                # Mapowanie typów zmian (jak w normalnym save)
                shift_type_mapping = {
                    'D': 'DNIOWKA',
                    'DNIOWKA': 'DNIOWKA',
                    'N': 'NOCKA', 
                    'NOCKA': 'NOCKA',
                    'P': 'POPOLUDNIOWKA'
                }
                
                if shift_type_raw in shift_type_mapping:
                    shift_type = shift_type_mapping[shift_type_raw]
                elif shift_type_raw and shift_type_raw.startswith('P '):
                    shift_type = shift_type_raw
                else:
                    shift_type = shift_type_raw
                
                if not date or not employee:
                    continue
                
                # Znajdź ID pracownika
                emp_row = db.execute("SELECT id FROM employees WHERE name = ?", (employee,)).fetchone()
                if not emp_row:
                    logger.warning(f"Nie znaleziono pracownika: {employee}")
                    continue
                
                emp_id = emp_row['id']
                
                # Zapisz tylko jeśli shift_type nie jest puste
                if shift_type and shift_type.strip() and shift_type != '-':
                    db.execute("""
                        INSERT INTO draft_shifts (date, shift_type, employee_id, created_by)
                        VALUES (?, ?, ?, ?)
                    """, (date, shift_type, emp_id, user_id))
                    logger.info(f"DRAFT ZAPISANO: {date} - {employee} - {shift_type}")
            
            db.commit()
            logger.info(f"Zapisano {len(changes)} zmian w trybie draft przez {user_email}")
            return jsonify(status="ok", message="Zmiany robocze zapisane pomyślnie")
        
        # Normalny tryb zapisywania (z powiadomieniami)
        # Zapisz zmiany do bazy danych
        for change in changes:
            date = change.get('date')
            employee = change.get('employee') or change.get('name')  # Support both formats
            shift_type_raw = change.get('shift_type') or change.get('value', '')  # Support both formats
            
            # Map abbreviated shift types to full names, but preserve custom text
            shift_type_mapping = {
                'D': 'DNIOWKA',
                'DNIOWKA': 'DNIOWKA',
                'N': 'NOCKA', 
                'NOCKA': 'NOCKA',
                'P': 'POPOLUDNIOWKA'
            }
            # If it's not a standard single-letter mapping, keep the original value (for custom text like "D 9-21", "D 10-22")
            if shift_type_raw in shift_type_mapping:
                shift_type = shift_type_mapping[shift_type_raw]
            elif shift_type_raw and shift_type_raw.startswith('P '):
                # Międzyzmiana w formacie "P 10-22" - zachowaj pełny tekst
                shift_type = shift_type_raw
            else:
                shift_type = shift_type_raw  # Keep custom text as-is
            
            if not date or not employee:
                continue
            
            # Znajdź ID pracownika
            emp_row = db.execute("SELECT id FROM employees WHERE name = ?", (employee,)).fetchone()
            if not emp_row:
                logger.warning(f"Nie znaleziono pracownika: {employee}")
                continue
            
            emp_id = emp_row['id']
            
            # Pobierz starą zmianę przed usunięciem
            old_shift = db.execute("SELECT shift_type FROM shifts WHERE date = ? AND employee_id = ?", 
                                 (date, emp_id)).fetchone()
            old_shift_type = old_shift['shift_type'] if old_shift else None
            
            # Usuń istniejącą zmianę dla tego pracownika w tym dniu
            db.execute("DELETE FROM shifts WHERE date = ? AND employee_id = ?", (date, emp_id))
            
            # Dodaj nową zmianę (jeśli nie jest pusta)
            if shift_type and shift_type.strip() and shift_type != '-':
                db.execute("INSERT INTO shifts (date, shift_type, employee_id) VALUES (?, ?, ?)", 
                          (date, shift_type, emp_id))
                logger.info(f"ZAPISANO: {date} - {employee} - {shift_type}")
                
                # Zapisz historię zmian
                action = "DODANO" if not old_shift_type else "ZMIENIONO"
                db.execute("""
                    INSERT INTO schedule_changes (date, employee_name, old_shift, new_shift, changed_by)
                    VALUES (?, ?, ?, ?, ?)
                """, (date, employee, old_shift_type or '', shift_type, user_email))
            else:
                # Zapisz usunięcie do schedule_changes
                if old_shift_type:
                    db.execute("""
                        INSERT INTO schedule_changes (date, employee_name, old_shift, new_shift, changed_by)
                        VALUES (?, ?, ?, ?, ?)
                    """, (date, employee, old_shift_type, '', user_email))
        
        # Użyj transakcji dla atomowości
        try:
            db.commit()
            
            # Invalidate cache for affected dates
            for change in changes:
                date = change.get('date')
                if date:
                    cache.delete(f"shifts:{date}")
            
            logger.info(f"Zapisano {len(changes)} zmian przez {user_email}")
            return jsonify(status="ok", message="Zmiany zapisane pomyślnie")
        except Exception as commit_error:
            db.rollback()
            logger.error(f"Błąd podczas commit: {commit_error}")
            return jsonify(error="Błąd podczas zapisywania zmian"), 500
        
    except Exception as e:
        logger.error(f"Błąd podczas zapisywania zmian: {e}")
        return jsonify(error="Wystąpił błąd podczas zapisywania"), 500

@bp.get("/slot")
@login_required
def api_slot():
    """Pobiera informacje o slocie"""
    try:
        date = request.args.get('date')
        employee = request.args.get('employee')
        is_draft = request.args.get('is_draft', 'false').lower() == 'true'
        
        if not date or not employee:
            return jsonify(error="Brak wymaganych parametrów"), 400
        
        db = get_db()
        
        # Znajdź ID pracownika
        emp_row = db.execute("SELECT id FROM employees WHERE name = ?", (employee,)).fetchone()
        if not emp_row:
            return jsonify(error="Nie znaleziono pracownika"), 404
        
        emp_id = emp_row['id']
        
        # Jeśli to tryb draft, sprawdź najpierw draft_shifts
        if is_draft:
            from ..database import ensure_draft_shifts_table
            ensure_draft_shifts_table()
            
            user_id = session.get('user_id')
            user_role = session.get('user_role', 'USER')
            
            # Tylko admini mogą używać trybu draft
            if user_role != 'ADMIN':
                return jsonify(error="Tylko administratorzy mogą używać trybu roboczego"), 403
            
            # Pobierz zmianę z draft dla tego pracownika w tym dniu
            draft_row = db.execute("""
                SELECT shift_type FROM draft_shifts 
                WHERE date = ? AND employee_id = ? AND created_by = ?
            """, (date, emp_id, user_id)).fetchone()
            
            shift_type = draft_row['shift_type'] if draft_row else ''
        else:
            # Normalny tryb - pobierz z głównej tabeli shifts
            shift_row = db.execute("""
                SELECT shift_type FROM shifts 
                WHERE date = ? AND employee_id = ?
            """, (date, emp_id)).fetchone()
            
            shift_type = shift_row['shift_type'] if shift_row else ''
        
        return jsonify({
            'date': date,
            'employee': employee,
            'shift_type': shift_type
        })
        
    except Exception as e:
        logger.error(f"Błąd podczas pobierania slota: {e}")
        return jsonify(error="Wystąpił błąd podczas pobierania danych"), 500

# ============================================================================
# ENDPOINTY SYSTEMU WYMIAN
# ============================================================================

def map_shift_types(shift_raw):
    """Mapuje surowe typy zmian na standardowe"""
    mapping = {
        'D': 'DNIOWKA',
        'N': 'NOCKA', 
        'P': 'POPOLUDNIOWKA',
        'DNIOWKA': 'DNIOWKA',
        'NOCKA': 'NOCKA',
        'POPOLUDNIOWKA': 'POPOLUDNIOWKA'
    }
    return mapping.get(shift_raw, shift_raw)

def validate_swap_request_data(data):
    """Waliduje dane prośby o zamianę"""
    required_fields = ['from_date', 'to_date', 'from_employee', 'to_employee']
    
    for field in required_fields:
        if not data.get(field):
            return False, f"Brak wymaganego pola: {field}"
    
    # Sprawdź czy daty są różne
    if data['from_date'] == data['to_date']:
        return False, "Data 'od' i 'do' nie może być taka sama"
    
    # Sprawdź czy pracownicy są różni
    if data['from_employee'] == data['to_employee']:
        return False, "Pracownik 'od' i 'do' nie może być ten sam"
    
    return True, "OK"

@bp.post("/swaps")
def api_swaps_create():
    # Sprawdź czy użytkownik jest zalogowany
    if not session.get('user_id'):
        return jsonify(error="Nie jesteś zalogowany"), 401
    """Tworzy prośbę o zamianę zmianami"""
    try:
        from ..database import ensure_swaps_table
        ensure_swaps_table()
        data = safe_get_json()
        
        # Waliduj dane prośby
        is_valid, error_msg = validate_swap_request_data(data)
        if not is_valid:
            return jsonify(error=error_msg), 400
        
        # Pobierz dane z requestu
        from_date = data['from_date']
        to_date = data['to_date']
        from_employee = data['from_employee']
        to_employee = data['to_employee']
        comment = data.get('comment', '')
        
        # Sprawdź czy użytkownik ma prawo do tworzenia tej prośby
        user_id = session.get('user_id')
        user_email = session.get('user_email', '')
        
        db = get_db()
        
        # Znajdź nazwę pracownika odpowiadającego użytkownikowi
        emp_row = db.execute("SELECT name FROM employees WHERE user_id = ?", (user_id,)).fetchone()
        employee_name = emp_row['name'] if emp_row else session.get('user_name', '')
        
        if from_employee != employee_name:
            return jsonify(error="Nie masz uprawnień do tworzenia tej prośby"), 403
        
        # Sprawdź czy pracownicy istnieją
        from_emp = db.execute("SELECT id FROM employees WHERE name = ?", (from_employee,)).fetchone()
        to_emp = db.execute("SELECT id FROM employees WHERE name = ?", (to_employee,)).fetchone()
        
        if not from_emp or not to_emp:
            return jsonify(error="Jeden z pracowników nie istnieje"), 400
        
        # Sprawdź czy nie ma już aktywnej prośby między tymi pracownikami
        existing = db.execute("""
            SELECT id FROM swap_requests 
            WHERE from_employee = ? AND to_employee = ? 
            AND recipient_status = 'PENDING'
        """, (from_employee, to_employee)).fetchone()
        
        if existing:
            return jsonify(error="Już istnieje aktywna prośba między tymi pracownikami"), 400
        
        # Utwórz prośbę
        user_id = session.get('user_id')
        cursor = db.execute("""
            INSERT INTO swap_requests 
            (requester_user_id, from_date, to_date, from_employee, to_employee, comment_requester)
            VALUES (?, ?, ?, ?, ?, ?)
        """, (user_id, from_date, to_date, from_employee, to_employee, comment))
        
        db.commit()
        
        logger.info(f"Utworzono prośbę o zamianę: {from_employee} -> {to_employee}")
        return jsonify(status="ok", message="Prośba o zamianę została utworzona")
        
    except Exception as e:
        logger.error(f"Błąd podczas tworzenia prośby o zamianę: {e}")
        return jsonify(error="Wystąpił błąd podczas tworzenia prośby"), 500

@bp.get("/swaps/inbox")
def api_swaps_inbox():
    # Sprawdź czy użytkownik jest zalogowany
    if not session.get('user_id'):
        return jsonify(error="Nie jesteś zalogowany"), 401
    """Pobiera prośby o zamianę dla użytkownika"""
    try:
        from ..database import ensure_swaps_table
        ensure_swaps_table()
        db = get_db()
        
        user_id = session.get('user_id')
        user_email = session.get('user_email', '')
        
        # Pobierz nazwę użytkownika
        user_row = db.execute("SELECT name FROM users WHERE id = ?", (user_id,)).fetchone()
        user_name = user_row['name'] if user_row else user_email
        
        # Znajdź nazwę pracownika odpowiadającego użytkownikowi
        emp_row = db.execute("SELECT name FROM employees WHERE user_id = ?", (user_id,)).fetchone()
        employee_name = emp_row['name'] if emp_row else user_name
        
        # Sprawdź czy użytkownik jest bossem
        user_role = session.get('user_role', 'USER')
        is_boss = user_role == 'ADMIN'
        
        if is_boss:
            # Boss widzi wszystkie prośby
            requests = db.execute("""
                SELECT * FROM swap_requests 
                ORDER BY created_at DESC
            """).fetchall()
        else:
            # Zwykły użytkownik widzi tylko swoje prośby (po nazwie pracownika)
            requests = db.execute("""
                SELECT * FROM swap_requests 
                WHERE from_employee = ? OR to_employee = ?
                ORDER BY created_at DESC
            """, (employee_name, employee_name)).fetchall()
        
        # Przygotuj dane do zwrócenia
        items = []
        for req in requests:
            item = {
                'id': req['id'],
                'from_date': req['from_date'],
                'to_date': req['to_date'],
                'from_employee': req['from_employee'],
                'to_employee': req['to_employee'],
                'comment_requester': req['comment_requester'],
                'recipient_status': req['recipient_status'],
                'boss_status': req['boss_status'],
                'created_at': req['created_at']
            }
            
            # Dodaj informacje o zmianach
            from_shift = db.execute("""
                SELECT shift_type FROM shifts s
                JOIN employees e ON s.employee_id = e.id
                WHERE s.date = ? AND e.name = ?
            """, (req['from_date'], req['from_employee'])).fetchone()
            
            to_shift = db.execute("""
                SELECT shift_type FROM shifts s
                JOIN employees e ON s.employee_id = e.id
                WHERE s.date = ? AND e.name = ?
            """, (req['to_date'], req['to_employee'])).fetchone()
            
            item['from_shift'] = from_shift['shift_type'] if from_shift else ''
            item['to_shift'] = to_shift['shift_type'] if to_shift else ''
            
            # Określ status końcowy
            if req['boss_status'] == 'APPROVED':
                item['final_status'] = 'ZATWIERDZONA'
            elif req['boss_status'] == 'REJECTED':
                item['final_status'] = 'ODRZUCONA'
            elif req['recipient_status'] == 'APPROVED':
                item['final_status'] = 'OCZEKUJACE'
            elif req['recipient_status'] == 'REJECTED':
                item['final_status'] = 'ODRZUCONA'
            else:
                item['final_status'] = 'OCZEKUJACE'
            
            items.append(item)
        
        return jsonify({
            'items': items,
            'is_boss': is_boss
        })
        
    except Exception as e:
        logger.error(f"Błąd podczas pobierania próśb o zamianę: {e}")
        return jsonify(error="Wystąpił błąd podczas pobierania danych"), 500

@bp.post("/unavailability")
@login_required
def api_unavailability_create():
    """Tworzy zgłoszenie niedyspozycji"""
    try:
        from ..database import ensure_unavailability_table
        ensure_unavailability_table()
        data = safe_get_json()
        
        month_year = data.get("month_year", "").strip()
        selected_days = data.get("selected_days", [])
        comment = data.get("comment", "").strip()
        
        if not month_year or not selected_days:
            return jsonify(error="Miesiąc i wybrane dni są wymagane"), 400
        
        # Pobierz dane użytkownika z sesji
        user_email = session.get('user_email', '')
        user_name = session.get('user_name', '')
        user_id = session.get('user_id', '')
        
        logger.info(f"Session data - user_email: {user_email}, user_name: {user_name}, user_id: {user_id}")
        
        if not user_email:
            return jsonify(error="Nie można określić użytkownika (brak email w sesji)"), 400
        
        db = get_db()
        
        # Znajdź ID pracownika po email (najlepsze rozwiązanie - email jest unikalny)
        emp_row = db.execute("SELECT id FROM employees WHERE email = ?", (user_email,)).fetchone()
        logger.info(f"Szukam pracownika po email '{user_email}': {emp_row}")
        
        if not emp_row:
            return jsonify(error=f"Nie znaleziono pracownika dla email='{user_email}'. Upewnij się, że Twoje konto jest powiązane z pracownikiem w systemie."), 400
        
        emp_id = emp_row['id']
        
        # Usunięto sprawdzenie duplikatów - można wysyłać nieograniczoną ilość zgłoszeń
        
        # Konwertuj wybrane dni na string
        selected_days_str = ','.join(selected_days)
        
        # Utwórz zgłoszenie
        db.execute("""
            INSERT INTO unavailability_requests (employee_id, month_year, comment, selected_days)
            VALUES (?, ?, ?, ?)
        """, (emp_id, month_year, comment or "Niedyspozycja", selected_days_str))
        
        db.commit()
        
        logger.info(f"Użytkownik {user_name} zgłosił niedyspozycję dla {month_year}: {selected_days_str}")
        return jsonify(message="Subskrypcja zapisana pomyślnie")
        
    except Exception as e:
        logger.error(f"Błąd podczas tworzenia zgłoszenia niedyspozycji: {e}")
        return jsonify(error="Wystąpił błąd podczas tworzenia zgłoszenia"), 500

@bp.get("/unavailability/inbox")
def api_unavailability_inbox():
    # Sprawdź czy użytkownik jest zalogowany
    if not session.get('user_id'):
        return jsonify(error="Nie jesteś zalogowany"), 401
    """Pobiera prośby o niedyspozycję dla użytkownika"""
    try:
        db = get_db()
        
        # Pobierz wszystkie zgłoszenia niedyspozycji z informacjami o pracownikach
        unavailability_requests = db.execute("""
            SELECT 
                ur.id,
                ur.employee_id,
                ur.month_year,
                ur.selected_days,
                ur.comment,
                ur.status,
                ur.boss_comment,
                ur.created_at,
                ur.updated_at,
                e.name as employee_name,
                e.email as employee_email
            FROM unavailability_requests ur
            JOIN employees e ON ur.employee_id = e.id
            ORDER BY ur.created_at DESC
        """).fetchall()
        
        # Konwertuj na listę słowników
        items = []
        for req in unavailability_requests:
            # Parsuj selected_days jeśli to JSON string
            selected_days = req['selected_days']
            if selected_days.startswith('[') and selected_days.endswith(']'):
                try:
                    import json
                    selected_days = json.loads(selected_days)
                except:
                    selected_days = selected_days.split(',') if selected_days else []
            else:
                selected_days = selected_days.split(',') if selected_days else []
            
            items.append({
                'id': req['id'],
                'employee_id': req['employee_id'],
                'employee_name': req['employee_name'],
                'employee_email': req['employee_email'],
                'month_year': req['month_year'],
                'selected_days': selected_days,
                'comment': req['comment'],
                'status': req['status'],
                'boss_comment': req['boss_comment'],
                'created_at': req['created_at'],
                'updated_at': req['updated_at']
            })
        
        # Sprawdź czy użytkownik jest szefem (można dodać logikę sprawdzania roli)
        is_boss = True  # Na razie wszyscy mogą widzieć wszystkie zgłoszenia
        
        return jsonify({
            'items': items,
            'is_boss': is_boss
        })
        
    except Exception as e:
        logger.error(f"Błąd podczas pobierania próśb o niedyspozycję: {e}")
        return jsonify(error="Wystąpił błąd podczas pobierania danych"), 500

@bp.post("/swaps/respond")
def api_swaps_respond():
    # Sprawdź czy użytkownik jest zalogowany
    if not session.get('user_id'):
        return jsonify(error="Nie jesteś zalogowany"), 401
    """Odpowiada na prośbę o zamianę"""
    try:
        data = safe_get_json()
        request_id = data.get('id')
        status = data.get('status')  # 'APPROVED' lub 'REJECTED'
        
        if not request_id or not status:
            return jsonify(error="Brak wymaganych parametrów"), 400
        
        if status not in ['APPROVED', 'REJECTED']:
            return jsonify(error="Nieprawidłowy status"), 400
        
        db = get_db()
        
        # Pobierz prośbę
        req = db.execute("SELECT * FROM swap_requests WHERE id = ?", (request_id,)).fetchone()
        if not req:
            return jsonify(error="Nie znaleziono prośby"), 404
        
        # Sprawdź czy użytkownik ma prawo do odpowiedzi
        user_id = session.get('user_id')
        emp_row = db.execute("SELECT name FROM employees WHERE user_id = ?", (user_id,)).fetchone()
        employee_name = emp_row['name'] if emp_row else session.get('user_name', '')
        
        if req['to_employee'] != employee_name:
            return jsonify(error="Nie masz uprawnień do odpowiedzi na tę prośbę"), 403
        
        # Zaktualizuj status
        db.execute("""
            UPDATE swap_requests 
            SET recipient_status = ? 
            WHERE id = ?
        """, (status, request_id))
        
        db.commit()
        
        logger.info(f"Użytkownik {user_name} {status} prośbę o zamianę {request_id}")
        return jsonify(status="ok", message=f"Prośba została {status}")
        
    except Exception as e:
        logger.error(f"Błąd podczas odpowiadania na prośbę: {e}")
        return jsonify(error="Wystąpił błąd podczas przetwarzania"), 500

@bp.post("/swaps/boss")
@admin_required
def api_swaps_boss():
    """Boss zatwierdza/odrzuca prośbę o zamianę"""
    try:
        data = safe_get_json()
        request_id = data.get('id')
        status = data.get('status')  # 'APPROVED' lub 'REJECTED'
        
        if not request_id or not status:
            return jsonify(error="Brak wymaganych parametrów"), 400
        
        if status not in ['APPROVED', 'REJECTED']:
            return jsonify(error="Nieprawidłowy status"), 400
        
        db = get_db()
        
        # Pobierz prośbę
        req = db.execute("SELECT * FROM swap_requests WHERE id = ?", (request_id)).fetchone()
        if not req:
            return jsonify(error="Nie znaleziono prośby"), 404
        
        # Zaktualizuj status bossa
        db.execute("""
            UPDATE swap_requests 
            SET boss_status = ? 
            WHERE id = ?
        """, (status, request_id))
        
        # Jeśli boss zatwierdził, wykonaj zamianę
        if status == 'APPROVED':
            # Pobierz dane o zmianach
            from_date = req['from_date']
            to_date = req['to_date']
            from_employee = req['from_employee']
            to_employee = req['to_employee']
            
            # Znajdź ID pracowników
            from_emp = db.execute("SELECT id FROM employees WHERE name = ?", (from_employee,)).fetchone()
            to_emp = db.execute("SELECT id FROM employees WHERE name = ?", (to_employee,)).fetchone()
            
            if from_emp and to_emp:
                from_emp_id = from_emp['id']
                to_emp_id = to_emp['id']
                
                # Pobierz istniejące zmiany
                from_shift = db.execute("""
                    SELECT shift_type FROM shifts 
                    WHERE date = ? AND employee_id = ?
                """, (from_date, from_emp_id)).fetchone()
                
                to_shift = db.execute("""
                    SELECT shift_type FROM shifts 
                    WHERE date = ? AND employee_id = ?
                """, (to_date, to_emp_id)).fetchone()
                
                from_shift_type = from_shift['shift_type'] if from_shift else ''
                to_shift_type = to_shift['shift_type'] if to_shift else ''
                
                # Usuń stare zmiany
                db.execute("DELETE FROM shifts WHERE date = ? AND employee_id = ?", (from_date, from_emp_id))
                db.execute("DELETE FROM shifts WHERE date = ? AND employee_id = ?", (to_date, to_emp_id))
                
                # Dodaj nowe zmiany (zamienione)
                if to_shift_type:
                    db.execute("INSERT INTO shifts (date, shift_type, employee_id) VALUES (?, ?, ?)", 
                              (from_date, to_shift_type, from_emp_id))
                
                if from_shift_type:
                    db.execute("INSERT INTO shifts (date, shift_type, employee_id) VALUES (?, ?, ?)", 
                              (to_date, from_shift_type, to_emp_id))
                
                # Zapisz historię zmian
                admin_email = session.get('user_email', 'admin')
                db.execute("""
                    INSERT INTO schedule_changes (date, employee_name, old_shift, new_shift, changed_by)
                    VALUES (?, ?, ?, ?, ?)
                """, (from_date, from_employee, from_shift_type, to_shift_type, admin_email))
                
                db.execute("""
                    INSERT INTO schedule_changes (date, employee_name, old_shift, new_shift, changed_by)
                    VALUES (?, ?, ?, ?, ?)
                """, (to_date, to_employee, to_shift_type, from_shift_type, admin_email))
        
        db.commit()
        
        logger.info(f"Admin {session.get('user_email')} {status} prośbę o zamianę {request_id}")
        return jsonify(status="ok", message=f"Prośba została {status}")
        
    except Exception as e:
        logger.error(f"Błąd podczas przetwarzania prośby przez bossa: {e}")
        return jsonify(error="Wystąpił błąd podczas przetwarzania"), 500

@bp.post("/swaps/clear")
@admin_required
def api_swaps_clear():
    """Czyści wszystkie prośby o zamianę (admin)"""
    try:
        from ..database import ensure_swaps_table
        ensure_swaps_table()
        db = get_db()
        
        db.execute("DELETE FROM swap_requests")
        db.commit()
        
        logger.info(f"Admin {session.get('user_email')} wyczyścił wszystkie prośby o zamianę")
        return jsonify(status="ok", message="Wszystkie prośby zostały wyczyszczone")
        
    except Exception as e:
        logger.error(f"Błąd podczas czyszczenia próśb: {e}")
        return jsonify(error="Wystąpił błąd podczas czyszczenia"), 500

# ============================================================================
# ENDPOINTY ZMIAN DZIENNYCH
# ============================================================================

@bp.get("/shifts/<date>")
@login_required
@monitor_performance
def api_shifts_for_date(date):
    """Pobiera zmiany dla konkretnej daty z cache"""
    try:
        import datetime as dt
        
        # Walidacja daty
        try:
            target_date = dt.datetime.strptime(date, '%Y-%m-%d').date()
        except ValueError:
            return jsonify(error="Nieprawidłowy format daty. Użyj YYYY-MM-DD"), 400
        
        # Check cache first
        cache_key = f"shifts:{date}"
        cached_result = cache.get(cache_key)
        if cached_result is not None:
            return jsonify(cached_result)
        
        db = get_db()
        
        # Pobierz zmiany dla tej daty
        shifts = db.execute("""
            SELECT s.shift_type, e.name 
            FROM shifts s 
            JOIN employees e ON s.employee_id = e.id 
            WHERE s.date = ?
            ORDER BY s.shift_type, e.name
        """, (date,)).fetchall()
        
        # Grupuj zmiany według typu
        shifts_by_type = {
            'DNIOWKA': [],
            'POPOLUDNIOWKA': [],
            'NOCKA': []
        }
        
        for shift in shifts:
            shift_type = shift['shift_type']
            name = shift['name']
            
            # Rozpoznaj typ zmiany na podstawie różnych formatów
            if shift_type == 'DNIOWKA' or shift_type == 'D':
                shifts_by_type['DNIOWKA'].append(name)
            elif (shift_type == 'POPOLUDNIOWKA' or shift_type == 'P' or 
                  (shift_type and shift_type.startswith('P '))):
                shifts_by_type['POPOLUDNIOWKA'].append(name)
            elif shift_type == 'NOCKA' or shift_type == 'N':
                shifts_by_type['NOCKA'].append(name)
        
        # Dodaj małe litery dla kompatybilności z frontendem
        result = {
            'date': date,
            'dniowka': shifts_by_type['DNIOWKA'],
            'popoludniowka': shifts_by_type['POPOLUDNIOWKA'],
            'nocka': shifts_by_type['NOCKA']
        }
        
        # Cache result for 5 minutes
        cache.set(cache_key, result, 300)
        
        return jsonify(result)
        
    except Exception as e:
        logger.error(f"Błąd podczas pobierania zmian dla daty {date}: {e}")
        return jsonify(error="Wystąpił błąd podczas pobierania danych"), 500

# ============================================================================
# ENDPOINTY ZARZĄDZANIA PRACOWNIKAMI
# ============================================================================

@bp.get("/employees")
@admin_required
def api_employees_list():
    """Pobiera listę pracowników (admin) z cache"""
    try:
        # Check cache first
        cache_key = "employees:list"
        cached_result = cache.get(cache_key)
        if cached_result is not None:
            return jsonify(cached_result)
        
        from ..database import ensure_employees_code_column, ensure_employees_email_column
        ensure_employees_code_column()  # Upewnij się, że kolumna code istnieje
        ensure_employees_email_column()  # Upewnij się, że kolumna email istnieje
        
        db = get_db()
        employees = db.execute("SELECT id, name, code, email FROM employees ORDER BY name").fetchall()
        
        employees_list = []
        for emp in employees:
            employees_list.append({
                'id': emp['id'],
                'name': emp['name'],
                'code': emp['code'] or '',  # Zwróć pusty string jeśli code jest None
                'email': emp['email'] or ''  # Zwróć pusty string jeśli email jest None
            })
        
        result = {"employees": employees_list}
        
        # Cache result for 10 minutes
        cache.set(cache_key, result, 600)
        
        return jsonify(result)
        
    except Exception as e:
        logger.error(f"Błąd podczas pobierania pracowników: {e}")
        return jsonify(error="Wystąpił błąd podczas pobierania danych"), 500

@bp.post("/employees")
@admin_required
def api_employees_create():
    """Dodaje nowego pracownika (admin)"""
    try:
        from ..database import ensure_employees_code_column, ensure_employees_email_column
        ensure_employees_code_column()  # Upewnij się, że kolumna code istnieje
        ensure_employees_email_column()  # Upewnij się, że kolumna email istnieje
        
        data = safe_get_json()
        name = data.get('name', '').strip()
        code = data.get('code', '').strip()
        email = data.get('email', '').strip()
        
        if not name:
            return jsonify(error="Nazwa pracownika jest wymagana"), 400
        
        db = get_db()
        
        # Sprawdź czy pracownik już istnieje
        existing = db.execute("SELECT id FROM employees WHERE name = ?", (name,)).fetchone()
        if existing:
            return jsonify(error="Pracownik o tej nazwie już istnieje"), 400
        
        # Sprawdź czy kod już istnieje (jeśli podano)
        if code:
            existing_code = db.execute("SELECT id FROM employees WHERE code = ?", (code,)).fetchone()
            if existing_code:
                return jsonify(error="Pracownik o tym kodzie już istnieje"), 400
        
        # Sprawdź czy email już istnieje (jeśli podano)
        if email:
            existing_email = db.execute("SELECT id FROM employees WHERE email = ?", (email,)).fetchone()
            if existing_email:
                return jsonify(error="Pracownik o tym emailu już istnieje"), 400
        
        # Dodaj pracownika
        cursor = db.execute("INSERT INTO employees (name, code, email) VALUES (?, ?, ?)", (name, code or None, email or None))
        db.commit()
        
        # Invalidate employees cache
        cache.delete("employees:list")
        
        # Jeśli podano email, dodaj go automatycznie do whitelisty
        if email:
            try:
                import os
                current_whitelist = os.environ.get('GOOGLE_WHITELIST', '').split(',')
                current_whitelist = [e.strip() for e in current_whitelist if e.strip()]
                
                if email not in current_whitelist:
                    current_whitelist.append(email)
                    new_whitelist = ','.join(current_whitelist)
                    os.environ['GOOGLE_WHITELIST'] = new_whitelist
                    logger.info(f"Email {email} został automatycznie dodany do whitelisty")
            except Exception as e:
                logger.warning(f"Nie udało się dodać emaila do whitelisty: {e}")
        
        logger.info(f"Admin {session.get('user_email')} dodał pracownika: {name} (kod: {code}, email: {email})")
        return jsonify(status="ok", message="Pracownik został dodany", id=cursor.lastrowid)
        
    except Exception as e:
        logger.error(f"Błąd podczas dodawania pracownika: {e}")
        return jsonify(error="Wystąpił błąd podczas dodawania"), 500

@bp.put("/employees/<int:emp_id>")
@admin_required
def api_employees_update(emp_id):
    """Aktualizuje pracownika (admin)"""
    try:
        from ..database import ensure_employees_code_column, ensure_employees_email_column
        ensure_employees_code_column()  # Upewnij się, że kolumna code istnieje
        ensure_employees_email_column()  # Upewnij się, że kolumna email istnieje
        
        data = safe_get_json()
        name = data.get('name', '').strip()
        code = data.get('code', '').strip()
        email = data.get('email', '').strip()
        
        if not name:
            return jsonify(error="Nazwa pracownika jest wymagana"), 400
        
        db = get_db()
        
        # Sprawdź czy pracownik istnieje
        existing = db.execute("SELECT id FROM employees WHERE id = ?", (emp_id,)).fetchone()
        if not existing:
            return jsonify(error="Nie znaleziono pracownika"), 404
        
        # Sprawdź czy nazwa nie jest już używana
        name_exists = db.execute("SELECT id FROM employees WHERE name = ? AND id != ?", (name, emp_id)).fetchone()
        if name_exists:
            return jsonify(error="Pracownik o tej nazwie już istnieje"), 400
        
        # Sprawdź czy kod nie jest już używany (jeśli podano)
        if code:
            code_exists = db.execute("SELECT id FROM employees WHERE code = ? AND id != ?", (code, emp_id)).fetchone()
            if code_exists:
                return jsonify(error="Pracownik o tym kodzie już istnieje"), 400
        
        # Sprawdź czy email nie jest już używany (jeśli podano)
        if email:
            email_exists = db.execute("SELECT id FROM employees WHERE email = ? AND id != ?", (email, emp_id)).fetchone()
            if email_exists:
                return jsonify(error="Pracownik o tym emailu już istnieje"), 400
        
        # Pobierz stary email przed aktualizacją
        old_emp = db.execute("SELECT email FROM employees WHERE id = ?", (emp_id,)).fetchone()
        old_email = old_emp['email'] if old_emp else None
        
        # Aktualizuj pracownika
        db.execute("UPDATE employees SET name = ?, code = ?, email = ? WHERE id = ?", (name, code or None, email or None, emp_id))
        db.commit()
        
        # Invalidate employees cache
        cache.delete("employees:list")
        
        # Synchronizuj whitelistę
        try:
            import os
            current_whitelist = os.environ.get('GOOGLE_WHITELIST', '').split(',')
            current_whitelist = [e.strip() for e in current_whitelist if e.strip()]
            
            # Usuń stary email z whitelisty (jeśli istniał)
            if old_email and old_email in current_whitelist:
                current_whitelist.remove(old_email)
                logger.info(f"Stary email {old_email} został usunięty z whitelisty")
            
            # Dodaj nowy email do whitelisty (jeśli podano)
            if email and email not in current_whitelist:
                current_whitelist.append(email)
                logger.info(f"Nowy email {email} został dodany do whitelisty")
            
            # Zaktualizuj zmienną środowiskową
            new_whitelist = ','.join(current_whitelist)
            os.environ['GOOGLE_WHITELIST'] = new_whitelist
            
        except Exception as e:
            logger.warning(f"Nie udało się zsynchronizować whitelisty: {e}")
        
        logger.info(f"Admin {session.get('user_email')} zaktualizował pracownika {emp_id}: {name} (kod: {code}, email: {email})")
        return jsonify(status="ok", message="Pracownik został zaktualizowany")
        
    except Exception as e:
        logger.error(f"Błąd podczas aktualizacji pracownika: {e}")
        return jsonify(error="Wystąpił błąd podczas aktualizacji"), 500

@bp.delete("/employees/<int:emp_id>")
@admin_required
def api_employees_delete(emp_id):
    """Usuwa pracownika (admin)"""
    try:
        db = get_db()
        
        # Sprawdź czy pracownik istnieje i pobierz email
        existing = db.execute("SELECT name, email FROM employees WHERE id = ?", (emp_id,)).fetchone()
        if not existing:
            return jsonify(error="Nie znaleziono pracownika"), 404
        
        # Pobierz email przed usunięciem
        emp_email = existing['email'] if existing['email'] else None
        
        # Usuń pracownika i jego zmiany
        db.execute("DELETE FROM shifts WHERE employee_id = ?", (emp_id,))
        
        db.execute("DELETE FROM employees WHERE id = ?", (emp_id,))
        db.commit()
        
        # Invalidate employees cache
        cache.delete("employees:list")
        
        # Usuń email z whitelisty (jeśli istniał)
        if emp_email:
            try:
                import os
                current_whitelist = os.environ.get('GOOGLE_WHITELIST', '').split(',')
                current_whitelist = [e.strip() for e in current_whitelist if e.strip()]
                
                if emp_email in current_whitelist:
                    current_whitelist.remove(emp_email)
                    new_whitelist = ','.join(current_whitelist)
                    os.environ['GOOGLE_WHITELIST'] = new_whitelist
                    logger.info(f"Email {emp_email} został usunięty z whitelisty")
            except Exception as e:
                logger.warning(f"Nie udało się usunąć emaila z whitelisty: {e}")
        
        logger.info(f"Admin {session.get('user_email')} usunął pracownika {emp_id}: {existing['name']}")
        return jsonify(status="ok", message="Pracownik został usunięty")
        
    except Exception as e:
        logger.error(f"Błąd podczas usuwania pracownika: {e}")
        return jsonify(error="Wystąpił błąd podczas usuwania"), 500

# ============================================================================
# ENDPOINTY ZARZĄDZANIA WHITELISTĄ
# ============================================================================

@bp.get("/whitelist")
@admin_required
def api_whitelist_get():
    """Pobiera aktualną whitelistę emaili (admin)"""
    try:
        import os
        whitelist = os.environ.get('GOOGLE_WHITELIST', '').split(',')
        whitelist = [email.strip() for email in whitelist if email.strip()]
        
        return jsonify({"emails": whitelist})
        
    except Exception as e:
        logger.error(f"Błąd podczas pobierania whitelisty: {e}")
        return jsonify(error="Wystąpił błąd podczas pobierania whitelisty"), 500

@bp.post("/whitelist")
@admin_required
def api_whitelist_add():
    """Dodaje email do whitelisty (admin)"""
    try:
        data = safe_get_json()
        email = data.get('email', '').strip().lower()
        
        if not email:
            return jsonify(error="Email jest wymagany"), 400
        
        # Sprawdź format emaila
        import re
        if not re.match(r'^[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}$', email):
            return jsonify(error="Nieprawidłowy format emaila"), 400
        
        import os
        current_whitelist = os.environ.get('GOOGLE_WHITELIST', '').split(',')
        current_whitelist = [e.strip() for e in current_whitelist if e.strip()]
        
        if email in current_whitelist:
            return jsonify(error="Email już jest na whitelistcie"), 400
        
        # Dodaj email do whitelisty
        current_whitelist.append(email)
        new_whitelist = ','.join(current_whitelist)
        
        # Zaktualizuj zmienną środowiskową (tylko dla tej sesji)
        os.environ['GOOGLE_WHITELIST'] = new_whitelist
        
        logger.info(f"Admin {session.get('user_email')} dodał email do whitelisty: {email}")
        return jsonify(status="ok", message="Email został dodany do whitelisty")
        
    except Exception as e:
        logger.error(f"Błąd podczas dodawania do whitelisty: {e}")
        return jsonify(error="Wystąpił błąd podczas dodawania do whitelisty"), 500

@bp.delete("/whitelist")
@admin_required
def api_whitelist_remove():
    """Usuwa email z whitelisty (admin)"""
    try:
        data = safe_get_json()
        email = data.get('email', '').strip().lower()
        
        if not email:
            return jsonify(error="Email jest wymagany"), 400
        
        import os
        current_whitelist = os.environ.get('GOOGLE_WHITELIST', '').split(',')
        current_whitelist = [e.strip() for e in current_whitelist if e.strip()]
        
        if email not in current_whitelist:
            return jsonify(error="Email nie jest na whitelistcie"), 400
        
        # Usuń email z whitelisty
        current_whitelist.remove(email)
        new_whitelist = ','.join(current_whitelist)
        
        # Zaktualizuj zmienną środowiskową (tylko dla tej sesji)
        os.environ['GOOGLE_WHITELIST'] = new_whitelist
        
        logger.info(f"Admin {session.get('user_email')} usunął email z whitelisty: {email}")
        return jsonify(status="ok", message="Email został usunięty z whitelisty")
        
    except Exception as e:
        logger.error(f"Błąd podczas usuwania z whitelisty: {e}")
        return jsonify(error="Wystąpił błąd podczas usuwania z whitelisty"), 500

# ============================================================================
# ENDPOINTY SYSTEMOWE
# ============================================================================

@bp.route("/healthz", methods=["GET"])
def healthz():
    """Health check endpoint"""
    import datetime
    return jsonify({
        "status": "ok",
        "timestamp": datetime.datetime.now().isoformat()
    })

@bp.get("/debug/env")
def debug_env():
    """Debug endpoint - informacje o środowisku (tylko w development)"""
    from flask import current_app
    if not current_app.config.get('DEBUG', False):
        return jsonify({"error": "Debug endpoint not available"}), 404
    
    return jsonify({
        "environment": "development",
        "has_client_id": bool(os.environ.get('GOOGLE_CLIENT_ID')),
        "has_client_secret": bool(os.environ.get('GOOGLE_CLIENT_SECRET')),
        "has_secret_key": bool(os.environ.get('SECRET_KEY'))
    })

@bp.get("/push/vapid-key")
def get_vapid_key():
    """Zwraca klucz publiczny VAPID dla frontendu"""
    try:
        from flask import current_app
        public_key = current_app.config.get('VAPID_PUBLIC_KEY')
        if not public_key:
            return jsonify(error="Brak konfiguracji VAPID"), 500
        
        return jsonify(public_key=public_key)
    except Exception as e:
        logger.error(f"Błąd pobierania klucza VAPID: {e}")
        return jsonify(error="Błąd serwera"), 500

@bp.get("/schedule/changes")
@login_required
def api_schedule_changes():
    """Pobiera ostatnie zmiany w grafiku dla powiadomień"""
    try:
        db = get_db()
        user_id = session.get("user_id")
        
        # Pobierz nazwę aktualnego użytkownika
        user_row = db.execute("SELECT name FROM users WHERE id=?", (user_id,)).fetchone()
        current_user_name = user_row["name"] if user_row else "Nieznany użytkownik"
        
        # Pobierz ostatnie zmiany z ostatnich 24 godzin
        changes = db.execute("""
            SELECT id, date, employee_name, old_shift, new_shift, changed_by, changed_at
            FROM schedule_changes 
            WHERE changed_at > datetime('now', '-1 day')
            ORDER BY changed_at DESC
            LIMIT 50
        """).fetchall()
        
        # Konwertuj na listę słowników
        changes_list = []
        for change in changes:
            changes_list.append({
                'id': change['id'],
                'date': change['date'],
                'employee_name': change['employee_name'],
                'old_shift': change['old_shift'],
                'new_shift': change['new_shift'],
                'changed_by': change['changed_by'],
                'changed_at': change['changed_at']
            })
        
        return jsonify({
            'changes': changes_list,
            'current_user_name': current_user_name
        })
        
    except Exception as e:
        logger.error(f"Błąd podczas pobierania zmian w grafiku: {e}")
        return jsonify(error="Wystąpił błąd podczas pobierania zmian"), 500

@bp.post("/push/subscribe")
@login_required
def subscribe_push():
    """Zapisuje subskrypcję push dla użytkownika"""
    try:
        subscription = request.get_json()
        user_id = session.get("user_id")
        
        logger.info(f"Push subscribe - user_id: {user_id}, subscription: {subscription}")
        
        if not subscription:
            logger.error("Brak danych subskrypcji")
            return jsonify(error="Brak danych subskrypcji"), 400
        
        if not user_id:
            logger.error("Brak user_id w sesji")
            return jsonify(error="Brak user_id w sesji"), 400
        
        # Walidacja danych subskrypcji
        required_fields = ['endpoint', 'keys']
        if not all(field in subscription for field in required_fields):
            return jsonify(error="Nieprawidłowe dane subskrypcji"), 400
        
        if 'keys' not in subscription or not subscription['keys']:
            return jsonify(error="Brak kluczy w subskrypcji"), 400
        
        # Import funkcji zapisywania subskrypcji
        from ..database import save_push_subscription
        if save_push_subscription(user_id, subscription):
            return jsonify(message="Subskrypcja zapisana pomyślnie")
        else:
            return jsonify(error="Błąd podczas zapisywania subskrypcji"), 500
            
    except Exception as e:
        logger.error(f"Błąd podczas zapisywania subskrypcji push: {e}")
        return jsonify(error="Błąd serwera"), 500

# ============================================================================
# ENDPOINTY SYSTEMU ROBOCZEGO (DRAFT)
# ============================================================================

@bp.post("/draft/save")
@login_required
@admin_required
def api_draft_save():
    """Zapisuje zmiany w trybie roboczym (tylko dla adminów)"""
    try:
        from ..database import ensure_draft_shifts_table
        ensure_draft_shifts_table()
        
        data = safe_get_json()
        changes = data.get('changes', [])
        
        if not changes:
            return jsonify(error="Brak zmian do zapisania"), 400
        
        db = get_db()
        user_id = session.get('user_id')
        user_email = session.get('user_email', 'nieznany')
        
        # Wyczyść wszystkie istniejące drafty dla tego użytkownika
        db.execute("DELETE FROM draft_shifts WHERE created_by = ?", (user_id,))
        
        # Zapisz nowe zmiany w trybie draft
        for change in changes:
            date = change.get('date')
            employee = change.get('employee') or change.get('name')
            shift_type_raw = change.get('shift_type') or change.get('value', '')
            
            # Mapowanie typów zmian (jak w normalnym save)
            shift_type_mapping = {
                'D': 'DNIOWKA',
                'DNIOWKA': 'DNIOWKA',
                'N': 'NOCKA', 
                'NOCKA': 'NOCKA',
                'P': 'POPOLUDNIOWKA'
            }
            
            if shift_type_raw in shift_type_mapping:
                shift_type = shift_type_mapping[shift_type_raw]
            elif shift_type_raw and shift_type_raw.startswith('P '):
                shift_type = shift_type_raw
            else:
                shift_type = shift_type_raw
            
            if not date or not employee:
                continue
            
            # Znajdź ID pracownika
            emp_row = db.execute("SELECT id FROM employees WHERE name = ?", (employee,)).fetchone()
            if not emp_row:
                logger.warning(f"Nie znaleziono pracownika: {employee}")
                continue
            
            emp_id = emp_row['id']
            
            # Zapisz zmianę (może być pusta jeśli usuwamy zmianę)
            # Pusta wartość oznacza że użytkownik usunął zmianę z oficjalnego grafiku
            db.execute("""
                INSERT INTO draft_shifts (date, shift_type, employee_id, created_by)
                VALUES (?, ?, ?, ?)
            """, (date, shift_type or '', emp_id, user_id))
            logger.info(f"DRAFT ZAPISANO: {date} - {employee} - {shift_type or 'PUSTE'}")
        
        db.commit()
        
        logger.info(f"Zapisano {len(changes)} zmian w trybie draft przez {user_email}")
        return jsonify(status="ok", message="Zmiany robocze zapisane pomyślnie")
        
    except Exception as e:
        logger.error(f"Błąd podczas zapisywania zmian draft: {e}")
        return jsonify(error="Wystąpił błąd podczas zapisywania"), 500

@bp.get("/draft/load")
@login_required
@admin_required
def api_draft_load():
    """Ładuje wersję roboczą (tylko dla adminów)"""
    try:
        from ..database import ensure_draft_shifts_table
        ensure_draft_shifts_table()
        
        db = get_db()
        user_id = session.get('user_id')
        
        # Pobierz wszystkie drafty dla tego użytkownika
        draft_shifts = db.execute("""
            SELECT ds.date, ds.shift_type, e.name as employee_name
            FROM draft_shifts ds
            JOIN employees e ON ds.employee_id = e.id
            WHERE ds.created_by = ?
            ORDER BY ds.date, e.name
        """, (user_id,)).fetchall()
        
        # Konwertuj na format podobny do normalnych zmian
        changes = []
        for shift in draft_shifts:
            changes.append({
                'date': shift['date'],
                'employee': shift['employee_name'],
                'shift_type': shift['shift_type']
            })
        
        logger.info(f"📥 [DRAFT] Ładowanie draft dla użytkownika {user_id}: znaleziono {len(changes)} zmian")
        
        return jsonify({
            'status': 'ok',
            'changes': changes,
            'count': len(changes)
        })
        
    except Exception as e:
        logger.error(f"Błąd podczas ładowania draft: {e}")
        return jsonify(error="Wystąpił błąd podczas ładowania"), 500

@bp.post("/draft/publish")
@login_required
@admin_required
def api_draft_publish():
    """Publikuje zmiany z draft do produkcji (tylko dla adminów)"""
    try:
        from ..database import ensure_draft_shifts_table, ensure_schedule_changes_table
        ensure_draft_shifts_table()
        ensure_schedule_changes_table()
        
        db = get_db()
        user_id = session.get('user_id')
        user_email = session.get('user_email', 'nieznany')
        
        # Pobierz wszystkie drafty dla tego użytkownika
        draft_shifts = db.execute("""
            SELECT ds.date, ds.shift_type, e.name as employee_name, e.id as employee_id
            FROM draft_shifts ds
            JOIN employees e ON ds.employee_id = e.id
            WHERE ds.created_by = ?
        """, (user_id,)).fetchall()
        
        if not draft_shifts:
            return jsonify(error="Brak zmian do publikacji"), 400
        
        # Znajdź wszystkie unikalne daty w draft
        draft_dates = list(set([shift['date'] for shift in draft_shifts]))
        logger.info(f"📅 Daty do zastąpienia w oficjalnym grafiku: {draft_dates}")
        
        # KROK 1: Wyczyść wszystkie zmiany w oficjalnym grafiku dla dat z draft
        for date in draft_dates:
            # Pobierz wszystkie istniejące zmiany dla tej daty (dla historii)
            existing_shifts = db.execute("""
                SELECT s.shift_type, e.name as employee_name
                FROM shifts s
                JOIN employees e ON s.employee_id = e.id
                WHERE s.date = ?
            """, (date,)).fetchall()
            
            # Usuń wszystkie zmiany dla tej daty
            deleted_count = db.execute("DELETE FROM shifts WHERE date = ?", (date,)).rowcount
            logger.info(f"🗑️ Usunięto {deleted_count} zmian z oficjalnego grafiku dla daty {date}")
            
            # Zapisz usunięcia do historii zmian
            for old_shift in existing_shifts:
                db.execute("""
                    INSERT INTO schedule_changes (date, employee_name, old_shift, new_shift, changed_by)
                    VALUES (?, ?, ?, ?, ?)
                """, (date, old_shift['employee_name'], old_shift['shift_type'], '', user_email))
        
        # KROK 2: Dodaj wszystkie zmiany z draft do oficjalnego grafiku
        for shift in draft_shifts:
            date = shift['date']
            employee_name = shift['employee_name']
            shift_type = shift['shift_type']
            employee_id = shift['employee_id']
            
            # Dodaj zmianę tylko jeśli nie jest pusta
            if shift_type and shift_type.strip() and shift_type != '-':
                db.execute("INSERT INTO shifts (date, shift_type, employee_id) VALUES (?, ?, ?)", 
                          (date, shift_type, employee_id))
                
                # Zapisz dodanie do historii zmian
                db.execute("""
                    INSERT INTO schedule_changes (date, employee_name, old_shift, new_shift, changed_by)
                    VALUES (?, ?, ?, ?, ?)
                """, (date, employee_name, '', shift_type, user_email))
                logger.info(f"✅ Dodano zmianę: {date} - {employee_name} - {shift_type}")
        
        # Usuń drafty po publikacji
        db.execute("DELETE FROM draft_shifts WHERE created_by = ?", (user_id,))
        
        db.commit()
        
        logger.info(f"🚀 Opublikowano {len(draft_shifts)} zmian z draft przez {user_email}")
        logger.info(f"📊 Zastąpiono oficjalny grafik dla {len(draft_dates)} dat: {draft_dates}")
        return jsonify(status="ok", message=f"Opublikowano {len(draft_shifts)} zmian dla {len(draft_dates)} dat")
        
    except Exception as e:
        logger.error(f"Błąd podczas publikacji draft: {e}")
        return jsonify(error="Wystąpił błąd podczas publikacji"), 500

@bp.post("/draft/discard")
@login_required
@admin_required
def api_draft_discard():
    """Usuwa wersję roboczą (tylko dla adminów)"""
    try:
        from ..database import ensure_draft_shifts_table
        ensure_draft_shifts_table()
        
        db = get_db()
        user_id = session.get('user_id')
        user_email = session.get('user_email', 'nieznany')
        
        # Pobierz liczbę draftów przed usunięciem
        count = db.execute("SELECT COUNT(*) as count FROM draft_shifts WHERE created_by = ?", (user_id,)).fetchone()['count']
        
        # Usuń wszystkie drafty dla tego użytkownika
        db.execute("DELETE FROM draft_shifts WHERE created_by = ?", (user_id,))
        db.commit()
        
        logger.info(f"Usunięto {count} zmian draft przez {user_email}")
        return jsonify(status="ok", message=f"Usunięto {count} zmian roboczych")
        
    except Exception as e:
        logger.error(f"Błąd podczas usuwania draft: {e}")
        return jsonify(error="Wystąpił błąd podczas usuwania"), 500

@bp.get("/draft/status")
@login_required
@admin_required
def api_draft_status():
    """Sprawdza status wersji roboczej (tylko dla adminów)"""
    try:
        from ..database import ensure_draft_shifts_table
        ensure_draft_shifts_table()
        
        db = get_db()
        user_id = session.get('user_id')
        
        # Sprawdź czy istnieją drafty dla tego użytkownika
        count = db.execute("SELECT COUNT(*) as count FROM draft_shifts WHERE created_by = ?", (user_id,)).fetchone()['count']
        
        # Pobierz datę ostatniej modyfikacji
        last_update = db.execute("""
            SELECT MAX(updated_at) as last_update 
            FROM draft_shifts 
            WHERE created_by = ?
        """, (user_id,)).fetchone()['last_update']
        
        return jsonify({
            'has_draft': count > 0,
            'count': count,
            'last_update': last_update
        })
        
    except Exception as e:
        logger.error(f"Błąd podczas sprawdzania statusu draft: {e}")
        return jsonify(error="Wystąpił błąd podczas sprawdzania statusu"), 500

# ============================================================================
# ENDPOINTY EKSPORTU
# ============================================================================

@bp.get("/export/excel")
@login_required
@admin_required
def api_export_excel():
    """Eksportuje grafik do pliku Excel - dokładnie jak na stronie"""
    try:
        from flask import send_file
        import tempfile
        import datetime as dt
        import calendar
        
        # Sprawdź czy openpyxl jest dostępne
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            from openpyxl.utils import get_column_letter
            OPENPYXL_AVAILABLE = True
        except ImportError:
            OPENPYXL_AVAILABLE = False
            logger.warning("openpyxl nie jest zainstalowane. Eksport do Excel będzie niedostępny.")
            return jsonify(error="Eksport do Excel nie jest dostępny. Zainstaluj openpyxl."), 500
        
        if not OPENPYXL_AVAILABLE:
            return jsonify(error="Eksport do Excel nie jest dostępny"), 500
        
        # Pobierz parametry
        year = int(request.args.get('year', 2025))
        month = int(request.args.get('month', 9))
        
        db = get_db()
        
        # Pobierz pracowników (tak jak na stronie)
        employees = db.execute("SELECT id, name FROM employees ORDER BY name").fetchall()
        
        # Przygotuj dane kalendarza (tak jak na stronie)
        def get_calendar_days(year, month):
            """Pobierz dni kalendarza dla danego miesiąca - tylko dni danego miesiąca od 1 do ostatniego"""
            day_names = ['Pon', 'Wt', 'Śr', 'Czw', 'Pt', 'Sob', 'Nie']
            
            first_day = dt.date(year, month, 1)
            last_day = dt.date(year, month, calendar.monthrange(year, month)[1])
            
            calendar_days = []
            current_date = first_day
            while current_date <= last_day:
                is_off = current_date.weekday() >= 5  # Sobota i niedziela
                is_today = current_date == dt.date.today()
                
                calendar_days.append({
                    "date": current_date.strftime('%Y-%m-%d'),
                    "iso": current_date.isoformat(),
                    "dd": current_date.strftime('%d'),
                    "abbr": day_names[current_date.weekday()],
                    "is_off": is_off,
                    "is_today": is_today
                })
                
                current_date += dt.timedelta(days=1)
            
            return calendar_days
        
        # Pobierz zmiany dla miesiąca (tak jak na stronie)
        def get_shifts_for_month(db, year, month):
            """Pobierz zmiany dla danego miesiąca"""
            start_date = f"{year:04d}-{month:02d}-01"
            if month == 12:
                end_date = f"{year + 1:04d}-01-01"
            else:
                end_date = f"{year:04d}-{month + 1:02d}-01"
            
            shifts = db.execute("""
                SELECT s.date, s.shift_type, e.name 
                FROM shifts s 
                JOIN employees e ON s.employee_id = e.id 
                WHERE s.date >= ? AND s.date < ?
                ORDER BY s.date, e.name
            """, (start_date, end_date)).fetchall()
            
            shifts_by_date = {}
            for shift in shifts:
                date = shift["date"]
                if date not in shifts_by_date:
                    shifts_by_date[date] = {}
                shifts_by_date[date][shift["name"]] = shift["shift_type"]
            
            return shifts_by_date
        
        # Pobierz dane
        calendar_days = get_calendar_days(year, month)
        shifts_by_date = get_shifts_for_month(db, year, month)
        
        # Przygotuj wiersze grafiku (tak jak na stronie)
        schedule_rows = []
        for day in calendar_days:
            row_data = {
                "date": day["date"],
                "iso": day["iso"],
                "dd": day["dd"],
                "abbr": day["abbr"],
                "is_off": day["is_off"],
                "is_today": day["is_today"],
                "shifts": shifts_by_date.get(day["date"], {})
            }
            schedule_rows.append(row_data)
        
        # Utwórz workbook
        wb = Workbook()
        ws = wb.active
        ws.title = f"Grafik {year}-{month:02d}"
        
        # Style
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="D32F2F", end_color="D32F2F", fill_type="solid")
        center_alignment = Alignment(horizontal="center", vertical="center")
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        off_day_fill = PatternFill(start_color="F5F5F5", end_color="F5F5F5", fill_type="solid")
        
        # Nagłówki - dokładnie jak na stronie
        headers = ["Data", "Dzień"]
        
        # Dodaj pracowników (z obsługą pary Ania i Bożena)
        pair = ['Ania', 'Bożena']
        pair_done = False
        employees_for_export = []
        for emp in employees:
            if emp['name'] in pair and not pair_done:
                headers.append("Ania i Bożena")
                employees_for_export.append("Ania i Bożena")
                pair_done = True
            elif emp['name'] in pair and pair_done:
                # Pomiń duplikat z pary
                continue
            elif emp['name'] == 'Maciej':
                # Pomiń Macieja
                continue
            else:
                headers.append(emp['name'])
                employees_for_export.append(emp['name'])
        
        # Dodaj kolumnę licznika
        headers.append("Licznik")
        
        # Zapisz nagłówki
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_alignment
            cell.border = border
        
        # Wygeneruj wiersze - dokładnie jak na stronie
        row = 2
        for day_row in schedule_rows:
            # Data
            ws.cell(row=row, column=1, value=day_row['dd'])
            ws.cell(row=row, column=1).border = border
            if day_row['is_off']:
                ws.cell(row=row, column=1).fill = off_day_fill
            
            # Dzień
            ws.cell(row=row, column=2, value=day_row['abbr'])
            ws.cell(row=row, column=2).border = border
            if day_row['is_off']:
                ws.cell(row=row, column=2).fill = off_day_fill
            
            # Zmiany pracowników - dokładnie jak na stronie
            col = 3
            for emp_name in employees_for_export:
                shift_value = ''
                
                if emp_name == "Ania i Bożena":
                    # Sprawdź czy któraś z pary ma zmianę
                    ania_shift = day_row['shifts'].get('Ania', '')
                    bozena_shift = day_row['shifts'].get('Bożena', '')
                    
                    if ania_shift == 'DNIOWKA' or bozena_shift == 'DNIOWKA':
                        shift_value = 'D'
                    elif ania_shift == 'NOCKA' or bozena_shift == 'NOCKA':
                        shift_value = 'N'
                    elif ania_shift == 'POPOLUDNIOWKA' or bozena_shift == 'POPOLUDNIOWKA':
                        shift_value = 'P'
                    elif ania_shift.startswith('P ') or bozena_shift.startswith('P '):
                        # Międzyzmiana z godzinami
                        if ania_shift.startswith('P '):
                            shift_value = ania_shift
                        else:
                            shift_value = bozena_shift
                else:
                    # Zwykły pracownik
                    shift_type = day_row['shifts'].get(emp_name, '')
                    if shift_type == 'DNIOWKA':
                        shift_value = 'D'
                    elif shift_type == 'NOCKA':
                        shift_value = 'N'
                    elif shift_type == 'POPOLUDNIOWKA':
                        shift_value = 'P'
                    elif shift_type.startswith('P '):
                        # Międzyzmiana z godzinami
                        shift_value = shift_type
                
                cell = ws.cell(row=row, column=col, value=shift_value)
                cell.border = border
                if day_row['is_off']:
                    cell.fill = off_day_fill
                
                col += 1
            
            # Licznik - dokładnie jak na stronie
            dniowka_count = len([s for s in day_row['shifts'].values() if s == 'DNIOWKA'])
            popoludniowka_count = len([s for s in day_row['shifts'].values() if s == 'POPOLUDNIOWKA'])
            nocka_count = len([s for s in day_row['shifts'].values() if s == 'NOCKA'])
            
            # Dla pary Ania i Bożena licz jako jedną osobę
            if 'Ania' in day_row['shifts'] and 'Bożena' in day_row['shifts']:
                if day_row['shifts']['Ania'] == 'DNIOWKA' and day_row['shifts']['Bożena'] == 'DNIOWKA':
                    dniowka_count -= 1  # Odejmij jedną, bo liczymy parę jako jedną
                elif day_row['shifts']['Ania'] == 'POPOLUDNIOWKA' and day_row['shifts']['Bożena'] == 'POPOLUDNIOWKA':
                    popoludniowka_count -= 1
                elif day_row['shifts']['Ania'] == 'NOCKA' and day_row['shifts']['Bożena'] == 'NOCKA':
                    nocka_count -= 1
            
            licznik_text = f"D:{dniowka_count} P:{popoludniowka_count} N:{nocka_count}"
            ws.cell(row=row, column=col, value=licznik_text)
            ws.cell(row=row, column=col).border = border
            if day_row['is_off']:
                ws.cell(row=row, column=col).fill = off_day_fill
            
            row += 1
        
        # Dostosuj szerokość kolumn
        for col in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(col)].width = 12
        
        # Zapisz do tymczasowego pliku
        with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as tmp_file:
            wb.save(tmp_file.name)
            tmp_file_path = tmp_file.name
        
        # Wyślij plik
        filename = f"grafik_sp4600_{year}_{month:02d}.xlsx"
        return send_file(
            tmp_file_path,
            as_attachment=True,
            download_name=filename,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        
    except Exception as e:
        logger.error(f"Błąd podczas eksportu do Excel: {e}")
        return jsonify(error="Wystąpił błąd podczas eksportu"), 500

