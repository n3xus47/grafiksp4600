"""
Funkcje administracyjne - eksport Excel, powiadomienia, niedyspozycje
Zachowuje pełną kompatybilność z istniejącymi funkcjami
"""

import os
import logging
from flask import Blueprint, request, jsonify, session, send_file
from ..auth import login_required, admin_required
from ..database import get_db

logger = logging.getLogger(__name__)

bp = Blueprint('admin', __name__)

# ============================================================================
# ENDPOINTY EKSPORTU
# ============================================================================

@bp.get("/export/excel")
@login_required
@admin_required
def api_export_excel():
    """Eksportuje grafik do pliku Excel"""
    try:
        # Sprawdź czy openpyxl jest dostępne
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            from openpyxl.utils import get_column_letter
            OPENPYXL_AVAILABLE = True
        except ImportError:
            OPENPYXL_AVAILABLE = False
            logger.warning("openpyxl nie jest zainstalowane. Eksport do Excel będzie niedostępny.")
            return jsonify(error="Eksport do Excel nie jest dostępny. Zainstaluj openpyxl."), 500
        
        if not OPENPYXL_AVAILABLE:
            return jsonify(error="Eksport do Excel nie jest dostępny"), 500
        
        # Pobierz parametry
        year = int(request.args.get('year', 2025))
        month = int(request.args.get('month', 9))
        
        db = get_db()
        
        # Pobierz pracowników
        employees = db.execute("SELECT id, name FROM employees ORDER BY name").fetchall()
        
        # Pobierz zmiany dla miesiąca
        start_date = f"{year:04d}-{month:02d}-01"
        if month == 12:
            end_date = f"{year + 1:04d}-01-01"
        else:
            end_date = f"{year:04d}-{month + 1:02d}-01"
        
        shifts = db.execute("""
            SELECT s.date, s.shift_type, e.name 
            FROM shifts s 
            JOIN employees e ON s.employee_id = e.id 
            WHERE s.date >= ? AND s.date < ?
            ORDER BY s.date, e.name
        """, (start_date, end_date)).fetchall()
        
        # Utwórz workbook
        wb = Workbook()
        ws = wb.active
        ws.title = f"Grafik {year}-{month:02d}"
        
        # Style
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="D32F2F", end_color="D32F2F", fill_type="solid")
        center_alignment = Alignment(horizontal="center", vertical="center")
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Nagłówki
        ws['A1'] = f"Grafik zmian pracowników - {year}-{month:02d}"
        ws['A1'].font = Font(bold=True, size=16)
        ws.merge_cells('A1:H1')
        
        # Nagłówki kolumn
        headers = ['Data', 'Dzień'] + [emp['name'] for emp in employees] + ['Dzień', 'Data']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_alignment
            cell.border = border
        
        # Przygotuj dane
        import datetime as dt
        import calendar
        
        first_day = dt.date(year, month, 1)
        last_day = dt.date(year, month, calendar.monthrange(year, month)[1])
        
        # Pobierz pierwszy poniedziałek tygodnia
        days_since_monday = first_day.weekday()
        start_date_obj = first_day - dt.timedelta(days=days_since_monday)
        
        # Wygeneruj 5 tygodni (35 dni)
        row = 4
        for i in range(35):
            current_date = start_date_obj + dt.timedelta(days=i)
            is_off = current_date.weekday() >= 5  # Sobota i niedziela
            
            # Data
            ws.cell(row=row, column=1, value=current_date.strftime('%Y-%m-%d'))
            ws.cell(row=row, column=1).border = border
            
            # Dzień
            ws.cell(row=row, column=2, value=current_date.strftime('%a'))
            ws.cell(row=row, column=2).border = border
            
            # Zmiany pracowników
            for col, emp in enumerate(employees, 3):
                shift_type = ''
                for shift in shifts:
                    if (shift['date'] == current_date.strftime('%Y-%m-%d') and 
                        shift['name'] == emp['name']):
                        shift_type = shift['shift_type']
                        break
                
                cell = ws.cell(row=row, column=col, value=shift_type)
                cell.border = border
                if is_off:
                    cell.fill = PatternFill(start_color="F5F5F5", end_color="F5F5F5", fill_type="solid")
            
            # Dzień (duplikat)
            ws.cell(row=row, column=len(employees) + 3, value=current_date.strftime('%a'))
            ws.cell(row=row, column=len(employees) + 3).border = border
            
            # Data (duplikat)
            ws.cell(row=row, column=len(employees) + 4, value=current_date.strftime('%Y-%m-%d'))
            ws.cell(row=row, column=len(employees) + 4).border = border
            
            row += 1
        
        # Dostosuj szerokość kolumn
        for col in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(col)].width = 12
        
        # Zapisz do tymczasowego pliku
        import tempfile
        with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as tmp_file:
            wb.save(tmp_file.name)
            tmp_file_path = tmp_file.name
        
        # Wyślij plik
        filename = f"grafik_sp4600_{year}_{month:02d}.xlsx"
        return send_file(
            tmp_file_path,
            as_attachment=True,
            download_name=filename,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        
    except Exception as e:
        logger.error(f"Błąd podczas eksportu do Excel: {e}")
        return jsonify(error="Wystąpił błąd podczas eksportu"), 500

# ============================================================================
# ENDPOINTY POWIADOMIEŃ
# ============================================================================

@bp.post("/push/send")
@login_required
def api_push_send():
    """Wysyła powiadomienie push do użytkownika"""
    try:
        data = request.get_json()
        title = data.get('title', '')
        body = data.get('body', '')
        
        if not title or not body:
            return jsonify(error="Tytuł i treść są wymagane"), 400
        
        # TODO: Implementacja wysyłania powiadomień push
        logger.info(f"Powiadomienie push: {title} - {body}")
        return jsonify(status="ok", message="Powiadomienie zostało wysłane")
        
    except Exception as e:
        logger.error(f"Błąd podczas wysyłania powiadomienia: {e}")
        return jsonify(error="Wystąpił błąd podczas wysyłania"), 500

@bp.post("/push/send-to-all")
@login_required
def api_push_send_to_all():
    """Wysyła powiadomienie push do wszystkich użytkowników (admin)"""
    try:
        # Sprawdź czy użytkownik jest administratorem
        db = get_db()
        user_id = session.get("user_id")
        if not user_id:
            return jsonify(error="Brak uprawnień"), 403
        
        user_row = db.execute("SELECT role FROM users WHERE id=?", (user_id,)).fetchone()
        if not user_row or user_row["role"] != "ADMIN":
            return jsonify(error="Brak uprawnień"), 403
        
        data = request.get_json()
        title = data.get('title', '')
        body = data.get('body', '')
        
        if not title or not body:
            return jsonify(error="Tytuł i treść są wymagane"), 400
        
        # TODO: Implementacja wysyłania do wszystkich
        logger.info(f"Powiadomienie push do wszystkich: {title} - {body}")
        return jsonify(status="ok", message="Powiadomienie zostało wysłane do wszystkich")
        
    except Exception as e:
        logger.error(f"Błąd podczas wysyłania powiadomienia do wszystkich: {e}")
        return jsonify(error="Wystąpił błąd podczas wysyłania"), 500

@bp.post("/push/cleanup")
@login_required
def api_push_cleanup():
    """Czyści nieaktywne subskrypcje push (admin)"""
    try:
        # Sprawdź czy użytkownik jest administratorem
        db = get_db()
        user_id = session.get("user_id")
        if not user_id:
            return jsonify(error="Brak uprawnień"), 403
        
        user_row = db.execute("SELECT role FROM users WHERE id=?", (user_id,)).fetchone()
        if not user_row or user_row["role"] != "ADMIN":
            return jsonify(error="Brak uprawnień"), 403
        
        # TODO: Implementacja czyszczenia subskrypcji
        logger.info(f"Admin {session.get('user_email')} wyczyścił subskrypcje push")
        return jsonify(status="ok", message="Subskrypcje zostały wyczyszczone")
        
    except Exception as e:
        logger.error(f"Błąd podczas czyszczenia subskrypcji: {e}")
        return jsonify(error="Wystąpił błąd podczas czyszczenia"), 500

# ============================================================================
# ENDPOINTY NIEDYSPOZYCJI
# ============================================================================

@bp.post("/unavailability")
@login_required
def api_unavailability_create():
    """Tworzy zgłoszenie niedyspozycji"""
    try:
        from ..database import ensure_unavailability_table
        ensure_unavailability_table()
        data = safe_get_json()
        
        month_year = data.get("month_year", "").strip()
        reason = data.get("reason", "").strip()
        
        if not month_year or not reason:
            return jsonify(error="Miesiąc i powód są wymagane"), 400
        
        # Pobierz nazwę użytkownika
        user_name = session.get('user_name', '')
        if not user_name:
            return jsonify(error="Nie można określić użytkownika"), 400
        
        db = get_db()
        
        # Znajdź ID pracownika
        emp_row = db.execute("SELECT id FROM employees WHERE name = ?", (user_name,)).fetchone()
        if not emp_row:
            return jsonify(error="Nie znaleziono pracownika"), 400
        
        emp_id = emp_row['id']
        
        # Sprawdź czy już istnieje zgłoszenie dla tego miesiąca
        existing = db.execute("""
            SELECT id FROM unavailability_requests 
            WHERE employee_id = ? AND month_year = ?
        """, (emp_id, month_year)).fetchone()
        
        if existing:
            return jsonify(error="Już istnieje zgłoszenie niedyspozycji dla tego miesiąca"), 400
        
        # Utwórz zgłoszenie
        db.execute("""
            INSERT INTO unavailability_requests (employee_id, month_year, reason)
            VALUES (?, ?, ?)
        """, (emp_id, month_year, reason))
        
        db.commit()
        
        logger.info(f"Użytkownik {user_name} zgłosił niedyspozycję dla {month_year}")
        return jsonify(status="ok", message="Zgłoszenie niedyspozycji zostało utworzone")
        
    except Exception as e:
        logger.error(f"Błąd podczas tworzenia zgłoszenia niedyspozycji: {e}")
        return jsonify(error="Wystąpił błąd podczas tworzenia zgłoszenia"), 500

@bp.get("/unavailability/inbox")
@login_required
def api_unavailability_inbox():
    """Pobiera zgłoszenia niedyspozycji"""
    try:
        from ..database import ensure_unavailability_table
        ensure_unavailability_table()
        db = get_db()
        
        # Pobierz wszystkie zgłoszenia
        requests = db.execute("""
            SELECT ur.*, e.name as employee_name
            FROM unavailability_requests ur
            JOIN employees e ON ur.employee_id = e.id
            ORDER BY ur.created_at DESC
        """).fetchall()
        
        requests_list = []
        for req in requests:
            requests_list.append({
                'id': req['id'],
                'employee_name': req['employee_name'],
                'month_year': req['month_year'],
                'reason': req['reason'],
                'status': req['status'],
                'created_at': req['created_at']
            })
        
        return jsonify(requests_list)
        
    except Exception as e:
        logger.error(f"Błąd podczas pobierania zgłoszeń niedyspozycji: {e}")
        return jsonify(error="Wystąpił błąd podczas pobierania danych"), 500

@bp.post("/unavailability/respond")
@admin_required
def api_unavailability_respond():
    """Odpowiada na zgłoszenie niedyspozycji (admin)"""
    try:
        data = safe_get_json()
        request_id = data.get('id')
        status = data.get('status')  # 'APPROVED' lub 'REJECTED'
        
        if not request_id or not status:
            return jsonify(error="Brak wymaganych parametrów"), 400
        
        if status not in ['APPROVED', 'REJECTED']:
            return jsonify(error="Nieprawidłowy status"), 400
        
        db = get_db()
        
        # Pobierz zgłoszenie
        req = db.execute("""
            SELECT ur.*, e.name as employee_name
            FROM unavailability_requests ur
            JOIN employees e ON ur.employee_id = e.id
            WHERE ur.id = ?
        """, (request_id,)).fetchone()
        
        if not req:
            return jsonify(error="Nie znaleziono zgłoszenia"), 404
        
        # Zaktualizuj status
        db.execute("""
            UPDATE unavailability_requests 
            SET status = ? 
            WHERE id = ?
        """, (status, request_id))
        
        # Jeśli zatwierdzono, dodaj zmiany "-" dla wszystkich dni w miesiącu
        if status == 'APPROVED':
            month_year = req['month_year']
            year, month = map(int, month_year.split('-'))
            
            # Pobierz wszystkie dni w miesiącu
            import datetime as dt
            import calendar
            
            first_day = dt.date(year, month, 1)
            last_day = dt.date(year, month, calendar.monthrange(year, month)[1])
            
            current_date = first_day
            while current_date <= last_day:
                date_str = current_date.strftime('%Y-%m-%d')
                
                # Usuń istniejące zmiany dla tego pracownika w tym dniu
                db.execute("DELETE FROM shifts WHERE date = ? AND employee_id = ?", 
                          (date_str, req['employee_id']))
                
                # Dodaj wpis z "-"
                db.execute("INSERT INTO shifts(date, shift_type, employee_id) VALUES (?, ?, ?)", 
                          (date_str, "-", req['employee_id']))
                
                current_date += dt.timedelta(days=1)
        
        db.commit()
        
        logger.info(f"Admin {session.get('user_email')} {status} zgłoszenie niedyspozycji {request_id}")
        return jsonify(status="ok", message=f"Zgłoszenie zostało {status}")
        
    except Exception as e:
        logger.error(f"Błąd podczas odpowiadania na zgłoszenie niedyspozycji: {e}")
        return jsonify(error="Wystąpił błąd podczas przetwarzania zgłoszenia"), 500

# ============================================================================
# ENDPOINTY DODATKOWE
# ============================================================================

@bp.get("/schedule/changes")
@login_required
def api_schedule_changes():
    """Pobiera zmiany w grafiku"""
    try:
        db = get_db()
        
        # Pobierz ostatnie zmiany
        changes = db.execute("""
            SELECT * FROM schedule_changes 
            ORDER BY created_at DESC 
            LIMIT 50
        """).fetchall()
        
        changes_list = []
        for change in changes:
            changes_list.append({
                'id': change['id'],
                'date': change['date'],
                'employee_name': change['employee_name'],
                'old_shift': change['old_shift'],
                'new_shift': change['new_shift'],
                'changed_by': change['changed_by'],
                'created_at': change['created_at']
            })
        
        return jsonify(changes_list)
        
    except Exception as e:
        logger.error(f"Błąd podczas pobierania zmian w grafiku: {e}")
        return jsonify(error="Wystąpił błąd podczas pobierania danych"), 500

@bp.post("/notifications/test")
@login_required
def api_notifications_test():
    """Test powiadomień"""
    try:
        logger.info(f"Test powiadomień przez {session.get('user_email')}")
        return jsonify(status="ok", message="Test powiadomień wykonany")
        
    except Exception as e:
        logger.error(f"Błąd podczas testu powiadomień: {e}")
        return jsonify(error="Wystąpił błąd podczas testu"), 500
