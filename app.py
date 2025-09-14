"""
Employee shift schedule management system
Google OAuth2 authentication system
"""

import os
import sqlite3
import datetime as dt
import secrets
import logging
from typing import Optional, Dict, Any, List
from zoneinfo import ZoneInfo
from datetime import timedelta
from functools import wraps

from flask import Flask, g, render_template, jsonify, request, redirect, url_for, session, abort, make_response, send_file, send_from_directory
from dotenv import load_dotenv
from authlib.integrations.flask_client import OAuth
import calendar
import json

# Konfiguracja logowania
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)

# Ładowanie zmiennych środowiskowych
load_dotenv()

# Web Push Notifications
try:
    from pywebpush import webpush, WebPushException
    WEB_PUSH_AVAILABLE = True
except ImportError:
    WEB_PUSH_AVAILABLE = False
    logger.warning("pywebpush nie jest zainstalowane. Web Push Notifications będą niedostępne.")

# Inicjalizacja aplikacji
app = Flask(__name__)

# Wybierz odpowiednią konfigurację na podstawie środowiska
from config import get_config
app.config.from_object(get_config())

# Obsługa proxy headers (HTTPS przez nginx)
from werkzeug.middleware.proxy_fix import ProxyFix
app.wsgi_app = ProxyFix(app.wsgi_app, x_for=1, x_proto=1, x_host=1, x_prefix=1)

# Dodatkowe nagłówki bezpieczeństwa
@app.after_request
def add_security_headers(response):
    """
    Dodaje nagłówki bezpieczeństwa do każdej odpowiedzi serwera.
    To jest funkcja, która uruchamia się po każdym żądaniu HTTP.
    """
    # Jeśli aplikacja działa na HTTPS (produkcja), dodaj nagłówki bezpieczeństwa
    if app.config.get('PREFERRED_URL_SCHEME') == 'https':
        # HSTS - zmusza przeglądarkę do używania tylko HTTPS przez rok
        hsts_max_age = app.config.get('HSTS_MAX_AGE', 31536000)  # 1 rok w sekundach
        response.headers['Strict-Transport-Security'] = f'max-age={hsts_max_age}; includeSubDomains; preload'
        
        # Zapobiega atakom MIME type sniffing
        response.headers['X-Content-Type-Options'] = 'nosniff'
        
        # Zapobiega ładowaniu strony w ramce (ochrona przed clickjacking)
        response.headers['X-Frame-Options'] = 'DENY'
        
        # Włącza ochronę przed XSS w starszych przeglądarkach
        response.headers['X-XSS-Protection'] = '1; mode=block'
        
        # Content Security Policy - kontroluje jakie zasoby może ładować strona
        csp_policy = app.config.get('CSP_POLICY', "upgrade-insecure-requests; block-all-mixed-content")
        response.headers["Content-Security-Policy"] = csp_policy
    
    # Wyłącz cache dla plików statycznych (CSS, JS) - zawsze pobieraj najnowsze wersje
    if request.path.startswith("/static/"):
        response.headers["Cache-Control"] = "no-cache, no-store, must-revalidate"
        response.headers["Pragma"] = "no-cache"
        response.headers["Expires"] = "0"
    
    return response
    

# Initialize OAuth (po załadowaniu zmiennych środowiskowych)
oauth = OAuth(app)

# Web Push Notifications - funkcje pomocnicze
def send_push_notification(subscription, title, body, data=None):
    """
    Wysyła powiadomienie push do subskrybenta
    """
    if not WEB_PUSH_AVAILABLE:
        logger.warning("Web Push nie jest dostępne - pywebpush nie zainstalowane")
        return False
    
    try:
        vapid_private_key = app.config.get('VAPID_PRIVATE_KEY')
        vapid_claim_email = app.config.get('VAPID_CLAIM_EMAIL')
        
        if not vapid_private_key or not vapid_claim_email:
            logger.error("Brak konfiguracji VAPID")
            return False
        
        notification_data = {
            'title': title,
            'body': body,
            'icon': '/static/PKN.WA.D-192.png',
            'badge': '/static/PKN.WA.D-192.png',
            'data': data or {}
        }
        
        webpush(
            subscription_info=subscription,
            data=json.dumps(notification_data),
            vapid_private_key=vapid_private_key,
            vapid_claims={"sub": f"mailto:{vapid_claim_email}"}
        )
        
        logger.info(f"Powiadomienie push wysłane: {title}")
        return True
        
    except WebPushException as e:
        logger.error(f"Błąd wysyłania powiadomienia push: {e}")
        return False
    except Exception as e:
        logger.error(f"Nieoczekiwany błąd podczas wysyłania powiadomienia: {e}")
        return False

def save_push_subscription(user_id, subscription):
    """
    Zapisuje subskrypcję push w bazie danych
    """
    try:
        logger.info(f"Zapisywanie subskrypcji push - user_id: {user_id}, subscription: {subscription}")
        
        conn = get_db()
        cursor = conn.cursor()
        
        # Sprawdź czy użytkownik już ma subskrypcję
        cursor.execute("SELECT id FROM push_subscriptions WHERE user_id = ?", (user_id,))
        existing = cursor.fetchone()
        logger.info(f"Istniejąca subskrypcja: {existing}")
        
        if existing:
            # Aktualizuj istniejącą subskrypcję
            logger.info("Aktualizuję istniejącą subskrypcję")
            cursor.execute("""
                UPDATE push_subscriptions 
                SET subscription_data = ?, updated_at = CURRENT_TIMESTAMP 
                WHERE user_id = ?
            """, (json.dumps(subscription), user_id))
        else:
            # Utwórz nową subskrypcję
            logger.info("Tworzę nową subskrypcję")
            cursor.execute("""
                INSERT INTO push_subscriptions (user_id, subscription_data, created_at, updated_at)
                VALUES (?, ?, CURRENT_TIMESTAMP, CURRENT_TIMESTAMP)
            """, (user_id, json.dumps(subscription)))
        
        conn.commit()
        conn.close()
        logger.info(f"Subskrypcja push zapisana dla użytkownika {user_id}")
        return True
        
    except Exception as e:
        logger.error(f"Błąd zapisywania subskrypcji push: {e}")
        logger.error(f"Szczegóły błędu: {type(e).__name__}: {str(e)}")
        return False

def get_push_subscriptions(user_id=None):
    """
    Pobiera subskrypcje push z bazy danych
    """
    try:
        conn = get_db()
        cursor = conn.cursor()
        
        if user_id:
            cursor.execute("SELECT subscription_data FROM push_subscriptions WHERE user_id = ?", (user_id,))
        else:
            cursor.execute("SELECT user_id, subscription_data FROM push_subscriptions")
        
        results = cursor.fetchall()
        conn.close()
        
        if user_id:
            return json.loads(results[0][0]) if results else None
        else:
            return [(row[0], json.loads(row[1])) for row in results]
            
    except Exception as e:
        logger.error(f"Błąd pobierania subskrypcji push: {e}")
        return None

def cleanup_expired_subscriptions():
    """
    Czyści wygasłe subskrypcje push z bazy danych
    """
    if not WEB_PUSH_AVAILABLE:
        return
    
    try:
        subscriptions = get_push_subscriptions()
        if not subscriptions:
            return
        
        expired_count = 0
        for user_id, subscription in subscriptions:
            try:
                # Próbuj wysłać testowe powiadomienie
                webpush(
                    subscription_info=subscription,
                    data=json.dumps({"test": True}),
                    vapid_private_key=app.config.get('VAPID_PRIVATE_KEY'),
                    vapid_claims={"sub": f"mailto:{app.config.get('VAPID_CLAIM_EMAIL')}"}
                )
            except WebPushException as e:
                if e.response and e.response.status_code in [410, 404]:
                    # Subskrypcja wygasła - usuń z bazy
                    conn = get_db()
                    cursor = conn.cursor()
                    cursor.execute("DELETE FROM push_subscriptions WHERE user_id = ?", (user_id,))
                    conn.commit()
                    conn.close()
                    expired_count += 1
                    logger.info(f"Usunięto wygasłą subskrypcję dla użytkownika {user_id}")
                else:
                    logger.warning(f"Błąd sprawdzania subskrypcji dla {user_id}: {e}")
        
        if expired_count > 0:
            logger.info(f"Usunięto {expired_count} wygasłych subskrypcji")
            
    except Exception as e:
        logger.error(f"Błąd czyszczenia wygasłych subskrypcji: {e}")

# Konfiguracja Google OAuth - lazy loading (ładowanie na żądanie)
def get_google_oauth():
    """
    Tworzy i zwraca konfigurację Google OAuth tylko raz.
    Lazy loading oznacza, że konfiguracja jest tworzona dopiero gdy jest potrzebna.
    """
    # Sprawdź czy już mamy skonfigurowane Google OAuth
    if not hasattr(get_google_oauth, '_google_oauth'):
        # Jeśli nie, to skonfiguruj Google OAuth
        get_google_oauth._google_oauth = oauth.register(
            name='google',  # Nazwa dostawcy OAuth
            client_id=os.environ.get("GOOGLE_CLIENT_ID"),  # ID aplikacji z Google Cloud Console
            client_secret=os.environ.get("GOOGLE_CLIENT_SECRET"),  # Sekret aplikacji z Google Cloud Console
            server_metadata_url='https://accounts.google.com/.well-known/openid-configuration',  # Adres konfiguracji Google
            client_kwargs={
                'scope': 'openid email profile'  # Jakie dane chcemy od Google (email i podstawowe info)
            }
        )
    return get_google_oauth._google_oauth

# Inicjalizacja Google OAuth
google = get_google_oauth()

# Inicjalizacja bazy danych
db_path = app.config.get('DATABASE_PATH', os.path.join(os.path.dirname(__file__), "app.db"))

def get_db():
    """
    Pobiera połączenie z bazą danych SQLite.
    Jeśli połączenie nie istnieje, tworzy nowe.
    Zwraca: obiekt połączenia sqlite3.Connection
    """
    # Sprawdź czy już mamy połączenie z bazą danych w obiekcie g (Flask global)
    if "db" not in g:
        try:
            # Połącz się z bazą danych SQLite
            g.db = sqlite3.connect(db_path)
            # Ustaw row_factory żeby wyniki były jak słowniki (łatwiejsze do użycia)
            g.db.row_factory = sqlite3.Row
            # Włącz sprawdzanie kluczy obcych (foreign keys) - ważne dla integralności danych
            g.db.execute("PRAGMA foreign_keys = ON")
        except sqlite3.Error as e:
            # Jeśli błąd połączenia, zapisz w logach i przerwij z błędem 500
            logger.error(f"Database connection error: {e}")
            abort(500, "Database error")
    return g.db

def close_db(e=None):
    """
    Zamyka połączenie z bazą danych.
    Ta funkcja jest wywoływana automatycznie przez Flask po zakończeniu żądania.
    Args:
        e: Obiekt wyjątku (nieużywany, dla Flask teardown)
    """
    # Pobierz połączenie z bazy danych z obiektu g i usuń je
    db = g.pop("db", None)
    if db is not None:
        try:
            # Zamknij połączenie z bazą danych
            db.close()
        except sqlite3.Error as e:
            # Jeśli wystąpi błąd przy zamykaniu, zapisz w logach
            logger.error(f"Błąd zamykania bazy danych: {e}")

app.teardown_appcontext(close_db)

# Dekoratory uwierzytelniania - funkcje które sprawdzają czy użytkownik może wykonać daną akcję
def login_required(view):
    """
    Dekorator który sprawdza czy użytkownik jest zalogowany.
    Jeśli nie, przekierowuje na stronę logowania.
    """
    @wraps(view)
    def wrapper(*args, **kwargs):
        # Sprawdź czy w sesji jest user_id (oznacza że użytkownik jest zalogowany)
        if not session.get("user_id"):
            return redirect(url_for("signin"))  # Przekieruj na stronę logowania
        return view(*args, **kwargs)  # Jeśli zalogowany, wykonaj funkcję
    return wrapper

def admin_required(view):
    """
    Dekorator który sprawdza czy użytkownik ma uprawnienia administratora.
    Jeśli nie, pokazuje błąd 403 (brak uprawnień).
    """
    @wraps(view)
    def wrapper(*args, **kwargs):
        # Najpierw sprawdź czy użytkownik jest zalogowany
        if not session.get("user_id"):
            return redirect(url_for("signin"))
        
        try:
            # Pobierz rolę użytkownika z bazy danych
            db = get_db()
            row = db.execute("SELECT role FROM users WHERE id=?", (session["user_id"],)).fetchone()
            role = row["role"] if row and row["role"] else "USER"  # Domyślnie USER jeśli nie ma roli
            
            # Sprawdź czy użytkownik ma rolę ADMIN
            if role != "ADMIN":
                logger.warning(f"Próba dostępu do funkcji admin przez użytkownika {session['user_id']}")
                abort(403, "Brak uprawnień administratora")
                
        except sqlite3.Error as e:
            logger.error(f"Błąd sprawdzania uprawnień: {e}")
            abort(500, "Błąd sprawdzania uprawnień")
            
        return view(*args, **kwargs)
    return wrapper

# Funkcje pomocnicze - małe funkcje które pomagają w różnych zadaniach
def load_whitelist() -> Optional[set]:
    """
    Ładuje listę dozwolonych emaili z pliku .env.
    Tylko te emaile mogą się zalogować do aplikacji.
    """
    raw = os.environ.get("WHITELIST_EMAILS", "").strip()
    if not raw:
        return None  # Jeśli nie ma listy, to wszyscy mogą się logować
    # Podziel po przecinkach, usuń spacje i zamień na małe litery
    return {e.strip().lower() for e in raw.split(",") if e.strip()}

def get_bool_field(row: sqlite3.Row, field_name: str, default: bool = False) -> bool:
    """
    Bezpiecznie pobiera pole boolean (prawda/fałsz) z wiersza bazy danych.
    Jeśli pole nie istnieje lub jest błędne, zwraca wartość domyślną.
    """
    try:
        return bool(row.get(field_name, default))
    except (KeyError, TypeError):
        return default  # Jeśli błąd, zwróć wartość domyślną

def validate_date_format(date_str: str) -> bool:
    """
    Sprawdza czy data ma poprawny format YYYY-MM-DD (np. 2024-01-15).
    Zwraca True jeśli format jest poprawny, False jeśli nie.
    """
    if not date_str:
        return False  # Pusty string to niepoprawna data
    try:
        dt.datetime.strptime(date_str, "%Y-%m-%d")  # Spróbuj sparsować datę
        return True
    except ValueError:
        return False  # Jeśli nie można sparsować, to zły format

def validate_shift_type(shift_type: str) -> bool:
    """
    Sprawdza czy typ zmiany jest poprawny.
    Akceptuje DNIOWKA, NOCKA oraz inne niestandardowe teksty (ale nie puste).
    """
    if not shift_type or not shift_type.strip():
        return False  # Pusty string to niepoprawny typ zmiany
    # Akceptuj standardowe typy oraz niestandardowe teksty (niepuste)
    return True

def safe_get_json() -> Dict[str, Any]:
    """
    Bezpiecznie pobiera dane JSON z żądania HTTP.
    Jeśli wystąpi błąd, zwraca pusty słownik zamiast crashować aplikację.
    """
    try:
        data = request.get_json(silent=True)  # Pobierz JSON, nie pokazuj błędów
        return data if data else {}  # Jeśli nie ma danych, zwróć pusty słownik
    except (ValueError, TypeError) as e:
        logger.error(f"Błąd parsowania JSON: {e}")
        return {}  # Zwróć pusty słownik zamiast crashować
    except Exception as e:
        logger.error(f"Nieoczekiwany błąd podczas parsowania JSON: {e}")
        return {}  # Zwróć pusty słownik zamiast crashować

# Rate limiting - ochrona przed zbyt dużą liczbą żądań od jednego użytkownika
from collections import defaultdict
import time

# Słownik przechowujący historię żądań dla każdego użytkownika
request_counts = defaultdict(list)

def check_rate_limit(identifier: str, max_requests: int = 100, window: int = 60) -> bool:
    """
    Sprawdza czy użytkownik nie przekroczył limitu żądań w danym czasie.
    identifier: identyfikator użytkownika (user_id lub IP)
    max_requests: maksymalna liczba żądań (domyślnie 100)
    window: okno czasowe w sekundach (domyślnie 60)
    """
    now = time.time()
    user_requests = request_counts[identifier]
    
    # Usuń stare żądania (starsze niż okno czasowe)
    user_requests[:] = [req_time for req_time in user_requests if now - req_time < window]
    
    # Sprawdź czy nie przekroczono limitu
    if len(user_requests) >= max_requests:
        return False  # Przekroczono limit
    
    # Dodaj nowe żądanie do listy
    user_requests.append(now)
    return True  # Wszystko OK

def rate_limit_required(max_requests: int = 100, window: int = 60):
    """
    Dekorator który ogranicza liczbę żądań od jednego użytkownika.
    Chroni przed atakami typu "spam" lub nadużyciami.
    """
    def decorator(view):
        @wraps(view)
        def wrapper(*args, **kwargs):
            # Użyj user_id jeśli użytkownik jest zalogowany, w przeciwnym razie IP
            identifier = session.get("user_id", request.remote_addr)
            if not check_rate_limit(identifier, max_requests, window):
                logger.warning(f"Rate limit przekroczony dla {identifier}")
                return jsonify(error="Zbyt wiele żądań, spróbuj ponownie później"), 429
            return view(*args, **kwargs)
        return wrapper
    return decorator

# --- Kalendarz świąt polskich --------------------------------------------------------
def _easter_sunday(year: int) -> dt.date:
    """
    Oblicza datę niedzieli wielkanocnej dla danego roku.
    Używa algorytmu Gregoriańskiego (anonimowy algorytm).
    To jest skomplikowana matematyka, ale działa poprawnie.
    """
    # Anonymous Gregorian algorithm - skomplikowana matematyka do obliczenia Wielkanocy
    a = year % 19
    b = year // 100
    c = year % 100
    d = b // 4
    e = b % 4
    f = (b + 8) // 25
    g = (b - f + 1) // 3
    h = (19 * a + b - d - g + 15) % 30
    i = c // 4
    k = c % 4
    l = (32 + 2 * e + 2 * i - h - k) % 7
    m = (a + 11 * h + 22 * l) // 451
    month = (h + l - 7 * m + 114) // 31
    day = ((h + l - 7 * m + 114) % 31) + 1
    return dt.date(year, month, day)


def polish_holidays(year: int) -> set:
    """
    Zwraca zbiór wszystkich świąt polskich w danym roku.
    Uwzględnia święta stałe (te same daty co roku) i ruchome (zależne od Wielkanocy).
    """
    # Oblicz datę Wielkanocy i święta z nią związane
    easter = _easter_sunday(year)
    easter_monday = easter + dt.timedelta(days=1)  # Poniedziałek Wielkanocny
    corpus_christi = easter + dt.timedelta(days=60)  # Boże Ciało

    # Święta stałe - te same daty co roku
    fixed = {
        dt.date(year, 1, 1),   # Nowy Rok
        dt.date(year, 1, 6),   # Trzech Króli
        dt.date(year, 5, 1),   # Święto Pracy
        dt.date(year, 5, 3),   # Święto Konstytucji 3 Maja
        dt.date(year, 8, 15),  # Wniebowzięcie NMP
        dt.date(year, 11, 1),  # Wszystkich Świętych
        dt.date(year, 11, 11), # Święto Niepodległości
        dt.date(year, 12, 25), # Boże Narodzenie (pierwszy dzień)
        dt.date(year, 12, 26), # Boże Narodzenie (drugi dzień)
    }

    # Święta ruchome - zależne od daty Wielkanocy
    moveable = {easter_monday, corpus_christi}
    return fixed | moveable | {easter}  # Połącz wszystkie święta w jeden zbiór


# --- Schemat bazy danych i proste polecenia CLI ---------------------------------------------------
@app.cli.command("init-db")
def init_db():
    """
    Tworzy tabele w bazie danych SQLite.
    To polecenie można uruchomić przez: flask init-db
    """
    db = get_db()
    db.executescript("""
    PRAGMA foreign_keys = ON;  -- Włącz sprawdzanie kluczy obcych

    -- Tabela użytkowników - przechowuje dane z Google OAuth
    CREATE TABLE IF NOT EXISTS users (
      id INTEGER PRIMARY KEY AUTOINCREMENT,
      google_sub TEXT NOT NULL UNIQUE,  -- Unikalny ID z Google
      email TEXT NOT NULL UNIQUE,       -- Email użytkownika
      name TEXT,                        -- Imię i nazwisko
      picture TEXT,                     -- URL do zdjęcia profilowego
      created_at TEXT DEFAULT (datetime('now'))  -- Data utworzenia konta
    );

    -- Tabela pracowników - lista osób które mogą mieć zmiany
    CREATE TABLE IF NOT EXISTS employees (
      id INTEGER PRIMARY KEY AUTOINCREMENT,
      name TEXT NOT NULL UNIQUE  -- Imię i nazwisko pracownika
    );

    -- Tabela zmian - grafik pracowników
    CREATE TABLE IF NOT EXISTS shifts (
      id INTEGER PRIMARY KEY AUTOINCREMENT,
      date TEXT NOT NULL,                    -- Data zmiany (YYYY-MM-DD)
      shift_type TEXT NOT NULL,              -- Typ zmiany (DNIOWKA, NOCKA, etc.)
      employee_id INTEGER NOT NULL REFERENCES employees(id) ON DELETE CASCADE,  -- Który pracownik
      UNIQUE(date, shift_type, employee_id)  -- Jeden pracownik może mieć tylko jedną zmianę danego typu w danym dniu
    );
    """)
    db.commit()
    print("db ready:", db_path)
# --- Swap requests -------------------------------------------------------------
# --- Funkcje migracji bazy danych --------------------------------------------------------
# Te funkcje dodają nowe kolumny do istniejących tabel bez usuwania danych
def ensure_users_role_column():
    """
    Dodaje kolumnę 'role' do tabeli users jeśli nie istnieje.
    To jest migracja - dodaje nową funkcjonalność do istniejącej bazy danych.
    """
    try:
        db = get_db()
        # Sprawdź jakie kolumny ma tabela users
        cols = db.execute("PRAGMA table_info(users)").fetchall()
        has_role = any(c[1] == 'role' for c in cols)
        
        if not has_role:
            # Dodaj kolumnę role z domyślną wartością 'USER'
            db.execute("ALTER TABLE users ADD COLUMN role TEXT DEFAULT 'USER'")
            db.commit()
            logger.info("Dodano kolumnę 'role' do tabeli users")
            
    except sqlite3.Error as e:
        logger.error(f"Błąd podczas dodawania kolumny role: {e}")
        db.rollback()
        raise

def ensure_employees_user_id_column():
    """
    Dodaje kolumnę 'user_id' do tabeli employees jeśli nie istnieje.
    Ta kolumna łączy pracowników z kontami użytkowników.
    """
    try:
        db = get_db()
        # Sprawdź jakie kolumny ma tabela employees
        cols = db.execute("PRAGMA table_info(employees)").fetchall()
        has_uid = any(c[1] == 'user_id' for c in cols)
        
        if not has_uid:
            # SQLite: nie można dodać kolumny z UNIQUE przez ALTER TABLE.
            # Dodajemy kolumnę bez ograniczenia, a unikalność wymuszamy indeksem.
            db.execute("ALTER TABLE employees ADD COLUMN user_id INTEGER REFERENCES users(id)")
            db.execute("CREATE UNIQUE INDEX IF NOT EXISTS employees_user_id_unique ON employees(user_id)")
            db.commit()
            logger.info("Dodano kolumnę 'user_id' do tabeli employees")
            
    except sqlite3.Error as e:
        logger.error(f"Błąd podczas dodawania kolumny user_id: {e}")
        db.rollback()
        raise

def ensure_employees_code_column():
    """Dodaje kolumnę code do tabeli employees jeśli nie istnieje"""
    try:
        db = get_db()
        cols = db.execute("PRAGMA table_info(employees)").fetchall()
        has_code = any(c[1] == 'code' for c in cols)
        
        if not has_code:
            # SQLite does not allow adding UNIQUE via ALTER TABLE ADD COLUMN reliably
            db.execute("ALTER TABLE employees ADD COLUMN code TEXT")
            db.execute("CREATE UNIQUE INDEX IF NOT EXISTS employees_code_idx ON employees(code)")
            db.commit()
            logger.info("Dodano kolumnę 'code' do tabeli employees")
            
    except sqlite3.Error as e:
        logger.error(f"Błąd podczas dodawania kolumny code: {e}")
        db.rollback()
        raise

def ensure_swaps_table():
    """Tworzy lub aktualizuje tabelę swap_requests"""
    try:
        db = get_db()
        
        # Sprawdź czy tabela istnieje
        table_exists = db.execute("""
            SELECT name FROM sqlite_master 
            WHERE type='table' AND name='swap_requests'
        """).fetchone()
        
        if not table_exists:
            # Utwórz nową tabelę
            db.executescript("""
        CREATE TABLE swap_requests (
          id INTEGER PRIMARY KEY AUTOINCREMENT,
          requester_user_id INTEGER NOT NULL,
          from_date TEXT NOT NULL,
          from_employee TEXT NOT NULL,
          to_date TEXT,  -- Allow NULL for give requests
          to_employee TEXT NOT NULL,
          comment_requester TEXT,
          recipient_status TEXT NOT NULL DEFAULT 'PENDING' CHECK(recipient_status IN ('PENDING','ACCEPTED','DECLINED')),
          recipient_comment TEXT,
          boss_status TEXT NOT NULL DEFAULT 'PENDING' CHECK(boss_status IN ('PENDING','APPROVED','REJECTED')),
          boss_comment TEXT,
          from_shift TEXT,
          to_shift TEXT,
          is_give_request BOOLEAN DEFAULT 0,
                  is_ask_request BOOLEAN DEFAULT 0,
          created_at TEXT DEFAULT (datetime('now')),
          updated_at TEXT DEFAULT (datetime('now'))
        );
                
                -- Indeksy dla lepszej wydajności
                CREATE INDEX IF NOT EXISTS idx_swap_from_date ON swap_requests(from_date);
                CREATE INDEX IF NOT EXISTS idx_swap_to_date ON swap_requests(to_date);
                CREATE INDEX IF NOT EXISTS idx_swap_requester ON swap_requests(requester_user_id);
                CREATE INDEX IF NOT EXISTS idx_swap_recipient_status ON swap_requests(recipient_status);
                CREATE INDEX IF NOT EXISTS idx_swap_boss_status ON swap_requests(boss_status);
            """)
            db.commit()
            logger.info("Utworzono tabelę swap_requests z indeksami")
            return
        
        # Sprawdź istniejące kolumny
        cols = db.execute("PRAGMA table_info(swap_requests)").fetchall()
        existing_columns = {c[1] for c in cols}
        
        # Dodaj brakujące kolumny
        columns_to_add = []
        
        if 'from_shift' not in existing_columns:
            columns_to_add.append("ALTER TABLE swap_requests ADD COLUMN from_shift TEXT")
        if 'to_shift' not in existing_columns:
            columns_to_add.append("ALTER TABLE swap_requests ADD COLUMN to_shift TEXT")
        if 'is_give_request' not in existing_columns:
            columns_to_add.append("ALTER TABLE swap_requests ADD COLUMN is_give_request BOOLEAN DEFAULT 0")
        if 'is_ask_request' not in existing_columns:
            columns_to_add.append("ALTER TABLE swap_requests ADD COLUMN is_ask_request BOOLEAN DEFAULT 0")
        if 'final_status' not in existing_columns:
            columns_to_add.append("ALTER TABLE swap_requests ADD COLUMN final_status TEXT DEFAULT 'PENDING'")
        
        # Wykonaj ALTER TABLE dla każdej brakującej kolumny
        for alter_statement in columns_to_add:
            try:
                db.execute(alter_statement)
                logger.info(f"Wykonano: {alter_statement}")
            except sqlite3.Error as e:
                logger.warning(f"Nie udało się wykonać {alter_statement}: {e}")
        
        # Sprawdź czy indeksy istnieją
        indexes = db.execute("PRAGMA index_list(swap_requests)").fetchall()
        existing_indexes = {idx[1] for idx in indexes}
        
        # Dodaj brakujące indeksy
        if 'idx_swap_from_date' not in existing_indexes:
            db.execute("CREATE INDEX idx_swap_from_date ON swap_requests(from_date)")
        if 'idx_swap_to_date' not in existing_indexes:
            db.execute("CREATE INDEX idx_swap_to_date ON swap_requests(to_date)")
        if 'idx_swap_requester' not in existing_indexes:
            db.execute("CREATE INDEX idx_swap_requester ON swap_requests(requester_user_id)")
        if 'idx_swap_recipient_status' not in existing_indexes:
            db.execute("CREATE INDEX idx_swap_recipient_status ON swap_requests(recipient_status)")
        if 'idx_swap_boss_status' not in existing_indexes:
            db.execute("CREATE INDEX idx_swap_boss_status ON swap_requests(boss_status)")
    
        db.commit()
        logger.info("Tabela swap_requests zaktualizowana")
        
    except sqlite3.Error as e:
        logger.error(f"Błąd podczas aktualizacji tabeli swap_requests: {e}")
        db.rollback()
        raise

def ensure_unavailability_table():
    """Tworzy lub aktualizuje tabelę unavailability_requests"""
    try:
        db = get_db()
        
        # Sprawdź czy tabela istnieje
        table_exists = db.execute("""
            SELECT name FROM sqlite_master 
            WHERE type='table' AND name='unavailability_requests'
        """).fetchone()
        
        if not table_exists:
            # Utwórz nową tabelę
            db.executescript("""
                CREATE TABLE unavailability_requests (
                  id INTEGER PRIMARY KEY AUTOINCREMENT,
                  employee_id INTEGER NOT NULL,
                  month_year TEXT NOT NULL,
                  selected_days TEXT NOT NULL,  -- JSON array of dates
                  comment TEXT,
                  status TEXT NOT NULL DEFAULT 'PENDING' CHECK(status IN ('PENDING','APPROVED','REJECTED')),
                  boss_comment TEXT,
                  created_at TEXT DEFAULT (datetime('now')),
                  updated_at TEXT DEFAULT (datetime('now')),
                  FOREIGN KEY (employee_id) REFERENCES employees(id) ON DELETE CASCADE
                );
                
                -- Indeksy dla lepszej wydajności
                CREATE INDEX IF NOT EXISTS idx_unavail_employee ON unavailability_requests(employee_id);
                CREATE INDEX IF NOT EXISTS idx_unavail_month ON unavailability_requests(month_year);
                CREATE INDEX IF NOT EXISTS idx_unavail_status ON unavailability_requests(status);
            """)
            db.commit()
            logger.info("Utworzono tabelę unavailability_requests z indeksami")
        else:
            logger.info("Tabela unavailability_requests już istnieje")
        
    except sqlite3.Error as e:
        logger.error(f"Błąd podczas tworzenia tabeli unavailability_requests: {e}")
        db.rollback()
        raise

def ensure_schedule_changes_table():
    """Tworzy tabelę schedule_changes jeśli nie istnieje"""
    try:
        db = get_db()
        
        # Sprawdź czy tabela istnieje
        table_exists = db.execute("""
            SELECT name FROM sqlite_master 
            WHERE type='table' AND name='schedule_changes'
        """).fetchone()
        
        if not table_exists:
            # Utwórz nową tabelę
            db.executescript("""
                CREATE TABLE schedule_changes (
                  id INTEGER PRIMARY KEY AUTOINCREMENT,
                  date TEXT NOT NULL,
                  employee_name TEXT NOT NULL,
                  shift_type TEXT,
                  action TEXT NOT NULL,
                  changed_by TEXT NOT NULL,
                  changed_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
                );
                CREATE INDEX IF NOT EXISTS idx_schedule_changes_date ON schedule_changes(date);
                CREATE INDEX IF NOT EXISTS idx_schedule_changes_employee ON schedule_changes(employee_name);
                CREATE INDEX IF NOT EXISTS idx_schedule_changes_changed_at ON schedule_changes(changed_at);
            """)
            db.commit()
            logger.info("Utworzono tabelę schedule_changes z indeksami")
        else:
            logger.info("Tabela schedule_changes już istnieje")
        
    except sqlite3.Error as e:
        logger.error(f"Błąd podczas tworzenia tabeli schedule_changes: {e}")
        db.rollback()

def ensure_push_subscriptions_table():
    """Tworzy tabelę push_subscriptions jeśli nie istnieje"""
    try:
        db = get_db()
        
        # Sprawdź czy tabela istnieje
        table_exists = db.execute("""
            SELECT name FROM sqlite_master 
            WHERE type='table' AND name='push_subscriptions'
        """).fetchone()
        
        if not table_exists:
            # Utwórz nową tabelę
            db.executescript("""
                CREATE TABLE push_subscriptions (
                  id INTEGER PRIMARY KEY AUTOINCREMENT,
                  user_id TEXT NOT NULL,
                  subscription_data TEXT NOT NULL,
                  created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                  updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
                );
                CREATE INDEX IF NOT EXISTS idx_push_subscriptions_user_id ON push_subscriptions(user_id);
                CREATE INDEX IF NOT EXISTS idx_push_subscriptions_created_at ON push_subscriptions(created_at);
            """)
            db.commit()
            logger.info("Utworzono tabelę push_subscriptions z indeksami")
        else:
            logger.info("Tabela push_subscriptions już istnieje")
        
    except sqlite3.Error as e:
        logger.error(f"Błąd podczas tworzenia tabeli push_subscriptions: {e}")
        db.rollback()
        raise

@app.post("/api/save")
@login_required
@rate_limit_required(max_requests=30, window=60)
def api_save():
    """Zapisuje zmiany w grafiku"""
    try:
        data = safe_get_json()
        changes = data.get("changes", [])
        
        if not changes:
            return jsonify(error="Brak zmian do zapisania"), 400
        
        db = get_db()
        ensure_schedule_changes_table()  # Upewnij się, że tabela istnieje
        logger.info(f"Przetwarzanie {len(changes)} zmian w api_save")
        
        # Pobierz nazwę użytkownika
        user_email = session.get('user_email', 'unknown')
        
        # Process each change
        processed_count = 0
        for i, change in enumerate(changes):
            logger.info(f"Zmiana #{i+1}: {change}")
            
            date = change.get("date")
            employee_name = change.get("employee") or change.get("name")  # Support both formats
            shift_type_raw = change.get("shift_type") or change.get("value")  # Support both formats
            
            # Map abbreviated shift types to full names, but preserve custom text
            shift_type_mapping = {
                'D': 'DNIOWKA',
                'DNIOWKA': 'DNIOWKA',
                'N': 'NOCKA', 
                'NOCKA': 'NOCKA',
                'P': 'POPOLUDNIOWKA'
            }
            # If it's not a standard single-letter mapping, keep the original value (for custom text like "D 9-21", "D 10-22")
            if shift_type_raw in shift_type_mapping:
                shift_type = shift_type_mapping[shift_type_raw]
            elif shift_type_raw and shift_type_raw.startswith('P '):
                # Międzyzmiana w formacie "P 10-22" - zachowaj pełny tekst
                shift_type = shift_type_raw
            else:
                shift_type = shift_type_raw  # Keep custom text as-is
            
            logger.info(f"Parsed - date: {date}, employee: {employee_name}, shift_raw: {shift_type_raw} -> shift: {shift_type}")
            
            if not all([date, employee_name]):
                logger.warning(f"Pominięto zmianę - brak date lub employee_name")
                continue
                
            # Get employee ID
            emp_row = db.execute("SELECT id FROM employees WHERE name=?", (employee_name,)).fetchone()
            if not emp_row:
                logger.warning(f"Nie znaleziono pracownika: {employee_name}")
                continue
                
            employee_id = emp_row["id"]
            logger.info(f"Employee ID: {employee_id}")
            
            # Pobierz poprzedni shift dla porównania
            old_shift = db.execute("SELECT shift_type FROM shifts WHERE date=? AND employee_id=?", (date, employee_id)).fetchone()
            old_shift_type = old_shift["shift_type"] if old_shift else None
            
            # Delete existing shift for this date/employee
            delete_result = db.execute("DELETE FROM shifts WHERE date=? AND employee_id=?", (date, employee_id))
            logger.info(f"DELETE query wykonane dla {date}, {employee_id}")
            
            # Add new shift if shift_type is not empty (accept any non-empty shift type)
            if shift_type and shift_type.strip():
                db.execute("INSERT INTO shifts(date, shift_type, employee_id) VALUES (?,?,?)", 
                          (date, shift_type, employee_id))
                logger.info(f"INSERT wykonane: {date}, {shift_type}, {employee_id}")
                processed_count += 1
                
                # Zapisz zmianę do schedule_changes
                action = "DODANO" if not old_shift_type else "ZMIENIONO"
                db.execute("""
                    INSERT INTO schedule_changes (date, employee_name, shift_type, action, changed_by)
                    VALUES (?, ?, ?, ?, ?)
                """, (date, employee_name, shift_type, action, user_email))
            else:
                logger.info(f"Brak shift_type lub pusty: {shift_type_raw} -> {shift_type}")
                
                # Zapisz usunięcie do schedule_changes
                if old_shift_type:
                    db.execute("""
                        INSERT INTO schedule_changes (date, employee_name, shift_type, action, changed_by)
                        VALUES (?, ?, ?, ?, ?)
                    """, (date, employee_name, old_shift_type, "USUNIETO", user_email))
        
        logger.info(f"Przetworzono {processed_count} zmian, wywołuję db.commit()")
        
        db.commit()
        logger.info(f"Zapisano {len(changes)} zmian w grafiku przez użytkownika {session.get('user_email')}")
        
        # Wyślij powiadomienia push o zmianach w grafiku
        try:
            send_schedule_change_notifications(changes, user_email)
        except Exception as e:
            logger.error(f"Błąd wysyłania powiadomień o zmianach: {e}")
        
        return jsonify(status="ok", message="Zmiany zostały zapisane")
        
    except Exception as e:
        logger.error(f"Błąd podczas zapisywania zmian: {e}")
        return jsonify(error="Wystąpił błąd podczas zapisywania"), 500

@app.get("/api/slot")
@login_required
def api_slot():
    date = request.args.get("date")
    name = request.args.get("employee")
    if not date or not name:
        return jsonify(error="missing"), 400
    db = get_db()
    emp = db.execute("SELECT id FROM employees WHERE name=?", (name,)).fetchone()
    st = None
    if emp:
        row = db.execute("SELECT shift_type FROM shifts WHERE date=? AND employee_id=?", (date, emp["id"]))
        row = row.fetchone()
        if row:
            st = row["shift_type"]
    return jsonify(shift_type=st)


@app.post("/api/swaps")
@login_required
@rate_limit_required(max_requests=20, window=60)  # Rate limiting dla swap requests
# Funkcje pomocnicze dla systemu wymian
def map_shift_types(shift_raw):
    """
    Mapuje skrócone typy zmian na pełne nazwy.
    """
    shift_type_mapping = {
        'D': 'DNIOWKA',
        'DNIOWKA': 'DNIOWKA',
        'N': 'NOCKA', 
        'NOCKA': 'NOCKA',
        'P': 'POPOLUDNIOWKA'
    }
    return shift_type_mapping.get(shift_raw, shift_raw) if shift_raw else None

def validate_swap_request_data(data):
    """
    Waliduje dane prośby o zamianę zmian.
    """
    from_employee = data.get("from_employee", "").strip()
    to_employee = data.get("to_employee", "").strip()
    comment = (data.get("comment") or "").strip()
    
    from_shift_raw = (data.get("from_shift") or "").strip().upper() or None
    to_shift_raw = (data.get("to_shift") or "").strip().upper() or None
    
    from_shift = map_shift_types(from_shift_raw)
    to_shift = map_shift_types(to_shift_raw)
    
    is_give_request = data.get("is_give_request", False)
    is_ask_request = data.get("is_ask_request", False)
    
    return {
        'from_employee': from_employee,
        'to_employee': to_employee,
        'comment': comment,
        'from_shift': from_shift,
        'to_shift': to_shift,
        'is_give_request': is_give_request,
        'is_ask_request': is_ask_request
    }

def api_swaps_create():
    """Tworzy nową prośbę o zamianę zmian"""
    try:
        ensure_swaps_table()
        data = safe_get_json()
        
        # Waliduj dane prośby
        request_data = validate_swap_request_data(data)
        from_employee = request_data['from_employee']
        to_employee = request_data['to_employee']
        comment = request_data['comment']
        from_shift = request_data['from_shift']
        to_shift = request_data['to_shift']
        is_give_request = request_data['is_give_request']
        is_ask_request = request_data['is_ask_request']

        # Walidacja podstawowych pól
        if not from_employee:
            return jsonify(error="Brak informacji o pracowniku oddającym zmianę"), 400
        if not to_employee:
            return jsonify(error="Brak informacji o pracowniku przejmującym zmianę"), 400
        
        # Walidacja typów zmian
        if from_shift and not validate_shift_type(from_shift):
            return jsonify(error="Nieprawidłowy typ zmiany oddawanej"), 400
        if to_shift and not validate_shift_type(to_shift):
            return jsonify(error="Nieprawidłowy typ zmiany przejmowanej"), 400
        
        # Walidacja dat
        from_date = data.get("from_date")
        to_date = data.get("to_date")
        
        # Dla ask requests, from_date może być null
        if not is_ask_request and from_date:
            if not validate_date_format(from_date):
                return jsonify(error="Nieprawidłowy format daty oddawanej zmiany"), 400
        
        # Dla give requests, to_date może być null
        if not is_give_request and to_date:
            if not validate_date_format(to_date):
                return jsonify(error="Nieprawidłowy format daty przejmowanej zmiany"), 400
        
        # Sprawdź czy nie próbuje zamienić z samym sobą
        if from_employee == to_employee:
            return jsonify(error="Nie możesz zamienić zmiany z samym sobą"), 400
        
        # Sprawdź czy użytkownik może prosić o zamianę swojej zmiany
        db = get_db()
        me_emp = db.execute("SELECT name FROM employees WHERE user_id=?", (session.get("user_id"),)).fetchone()
        me_name = (me_emp["name"] if me_emp else (session.get("user_name") or "")).strip()
        
        if me_name != from_employee:
            logger.warning(f"Użytkownik {me_name} próbował prosić o zamianę zmiany {from_employee}")
            return jsonify(error="Możesz prosić tylko o zamianę swoich zmian"), 403
        
        # Walidacja że requester rzeczywiście ma zmianę na from_date (tylko dla give requests i regular swaps)
        if from_date and not is_ask_request:
            from_emp_id = db.execute("SELECT id FROM employees WHERE name=?", (from_employee,)).fetchone()
            if from_emp_id:
                from_shift_exists = db.execute(
                    "SELECT shift_type FROM shifts WHERE date=? AND employee_id=?", 
                    (from_date, from_emp_id["id"])
                ).fetchone()
                if not from_shift_exists:
                    return jsonify(error=f"Nie masz zmiany w dniu {from_date}"), 400
    
        # Walidacja że target person ma zmianę na to_date (tylko dla regular swaps)
        if not is_give_request and to_date:
            to_emp_id = db.execute("SELECT id FROM employees WHERE name=?", (to_employee,)).fetchone()
            if to_emp_id:
                to_shift_exists = db.execute(
                    "SELECT shift_type FROM shifts WHERE date=? AND employee_id=?", 
                    (to_date, to_emp_id["id"])
                ).fetchone()
                if not to_shift_exists:
                    return jsonify(error=f"{to_employee} nie ma zmiany w dniu {to_date}"), 400
    
        # Sprawdź czy daty nie są już zaangażowane w pending swap requests
        existing_conflicts = check_existing_conflicts(db, from_date, from_employee, to_date, to_employee, 
                                                   is_give_request, is_ask_request)
        if existing_conflicts:
            return jsonify(error=". ".join(existing_conflicts)), 400
        
        # Zapisz prośbę o zamianę
        db.execute("""
            INSERT INTO swap_requests(
                requester_user_id, from_date, from_employee, to_date, to_employee, 
                comment_requester, from_shift, to_shift, is_give_request, is_ask_request, final_status
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """, (
            session["user_id"], from_date, from_employee, to_date, to_employee, 
            comment, from_shift, to_shift, is_give_request, is_ask_request, 'OCZEKUJACE'
        ))
        db.commit()
        
        logger.info(f"Utworzono prośbę o zamianę: {from_employee} -> {to_employee}")
        return jsonify(status="ok", message="Prośba o zamianę została utworzona")
        
    except sqlite3.Error as e:
        logger.error(f"Błąd bazy danych podczas tworzenia prośby o zamianę: {e}")
        return jsonify(error="Błąd bazy danych"), 500
    except ValueError as e:
        logger.error(f"Błąd walidacji danych: {e}")
        return jsonify(error="Nieprawidłowe dane"), 400
    except Exception as e:
        logger.error(f"Nieoczekiwany błąd podczas tworzenia prośby o zamianę: {e}")
        return jsonify(error="Wystąpił nieoczekiwany błąd"), 500

def _check_request_conflicts(req, from_date, from_employee, to_date, to_employee, is_give_request, is_ask_request):
    """Sprawdza konflikty dla pojedynczego requestu"""
    conflicts = []
    
    if is_give_request:
        if req["from_date"] == from_date and req["from_employee"] == from_employee:
            conflicts.append(f"Twoja zmiana w dniu {from_date} jest już zaangażowana w prośbę o zamianę")
        elif req["to_date"] == from_date and req["to_employee"] == from_employee:
            conflicts.append(f"Twoja zmiana w dniu {from_date} jest już zaangażowana w prośbę o zamianę")
    elif is_ask_request:
        if req["to_date"] == to_date and req["to_employee"] == to_employee:
            conflicts.append(f"Zmiana {to_employee} w dniu {to_date} jest już zaangażowana w prośbę o zamianę")
    else:
        # Regular swap - sprawdź obie daty
        if req["from_date"] == from_date and req["from_employee"] == from_employee:
            conflicts.append(f"Twoja zmiana w dniu {from_date} jest już zaangażowana w prośbę o zamianę")
        elif req["from_date"] == to_date and req["from_employee"] == to_employee:
            conflicts.append(f"Zmiana {to_employee} w dniu {to_date} jest już zaangażowana w prośbę o zamianę")
        elif req["to_date"] == from_date and req["to_employee"] == from_employee:
            conflicts.append(f"Twoja zmiana w dniu {from_date} jest już zaangażowana w prośbę o zamianę")
        elif req["to_date"] == to_date and req["to_employee"] == to_employee:
            conflicts.append(f"Zmiana {to_employee} w dniu {to_date} jest już zaangażowana w prośbę o zamianę")
    
    return conflicts

def check_existing_conflicts(db, from_date, from_employee, to_date, to_employee, is_give_request, is_ask_request):
    """Sprawdza czy daty nie są już zaangażowane w inne prośby o zamianę"""
    conflicts = []
    
    try:
        if is_give_request:
            # Dla give requests, sprawdź tylko from_date conflicts
            existing_requests = db.execute("""
                SELECT * FROM swap_requests 
                WHERE recipient_status IN ('PENDING', 'ACCEPTED') 
                AND boss_status IN ('PENDING', 'APPROVED')
                AND (
                    (from_date = ? AND from_employee = ?) OR
                    (to_date = ? AND to_employee = ?)
                )
            """, (from_date, from_employee, from_date, to_employee)).fetchall()
            
        elif is_ask_request:
            # Dla ask requests, sprawdź tylko to_date conflicts
            existing_requests = db.execute("""
                SELECT * FROM swap_requests 
                WHERE recipient_status IN ('PENDING', 'ACCEPTED') 
                AND boss_status IN ('PENDING', 'APPROVED')
                AND (to_date = ? AND to_employee = ?)
            """, (to_date, to_employee)).fetchall()
            
        else:
            # Dla regular swaps, sprawdź obie daty
            existing_requests = db.execute("""
                SELECT * FROM swap_requests 
                WHERE recipient_status IN ('PENDING', 'ACCEPTED') 
                AND boss_status IN ('PENDING', 'APPROVED')
                AND (
                    (from_date = ? AND from_employee = ?) OR
                    (to_date = ? AND to_employee = ?)
                )
                """, (from_date, from_employee, to_date, to_employee)).fetchall()
    
        # Znajdź konkretne konflikty
        for req in existing_requests:
            conflicts.extend(_check_request_conflicts(req, from_date, from_employee, to_date, to_employee, is_give_request, is_ask_request))
        
        return conflicts
        
    except sqlite3.Error as e:
        logger.error(f"Błąd podczas sprawdzania konfliktów: {e}")
        return [f"Błąd bazy danych: {e}"]

def calculate_final_status(recipient_status, boss_status):
    """Oblicza finalny status prośby na podstawie statusów pracownika i szefa"""
    if recipient_status == 'DECLINED':
        return 'ODRZUCONE'
    elif recipient_status == 'ACCEPTED' and boss_status == 'APPROVED':
        return 'ZATWIERDZONE'
    elif recipient_status == 'ACCEPTED' and boss_status == 'REJECTED':
        return 'ODRZUCONE_PRZEZ_SZEFA'
    elif recipient_status == 'ACCEPTED' and boss_status == 'PENDING':
        return 'WSTEPNIE_ZATWIERDZONE'
    else:
        return 'OCZEKUJACE'

def process_regular_swap(db, from_emp_id, to_emp_id, from_date, to_date, from_employee, to_employee):
    """Przetwarza regularną zamianę zmian"""
    try:
        logger.info(f"process_regular_swap: {from_employee}({from_emp_id}) <-> {to_employee}({to_emp_id}) na daty {from_date} <-> {to_date}")
        
        # Pobierz obecne zmiany obu pracowników
        from_shift = db.execute("SELECT shift_type FROM shifts WHERE date=? AND employee_id=?", (from_date, from_emp_id)).fetchone()
        to_shift = db.execute("SELECT shift_type FROM shifts WHERE date=? AND employee_id=?", (to_date, to_emp_id)).fetchone()
        
        logger.info(f"from_shift: {from_shift}, to_shift: {to_shift}")
        
        if not from_shift:
            logger.error(f"{from_employee} nie ma zmiany w dniu {from_date}")
            return False, f"{from_employee} nie ma zmiany w dniu {from_date}"
        if not to_shift:
            logger.error(f"{to_employee} nie ma zmiany w dniu {to_date}")
            return False, f"{to_employee} nie ma zmiany w dniu {to_date}"
        
        from_shift_type = from_shift["shift_type"]
        to_shift_type = to_shift["shift_type"]
        
        logger.info(f"from_shift_type: {from_shift_type}, to_shift_type: {to_shift_type}")
        
        # Wykonaj zamianę: pracownicy wymieniają się datami (nie typami zmian)
        # Usuń istniejące zmiany
        db.execute("DELETE FROM shifts WHERE date=? AND employee_id=?", (from_date, from_emp_id))
        db.execute("DELETE FROM shifts WHERE date=? AND employee_id=?", (to_date, to_emp_id))
        
        # Przypisz zamienione zmiany
        # from_employee dostaje zmianę to_employee na to_date
        db.execute("INSERT INTO shifts(date, shift_type, employee_id) VALUES (?,?,?)", (to_date, to_shift_type, from_emp_id))
        
        # to_employee dostaje zmianę from_employee na from_date
        db.execute("INSERT INTO shifts(date, shift_type, employee_id) VALUES (?,?,?)", (from_date, from_shift_type, to_emp_id))
        
        logger.info(f"Zamiana wykonana: {from_employee} <-> {to_employee}")
        return True, None
        
    except sqlite3.Error as e:
        logger.error(f"Błąd podczas przetwarzania regular swap: {e}")
        return False, f"Błąd bazy danych: {e}"

@app.get("/api/swaps/inbox")
@login_required
def api_swaps_inbox():
    """Pobiera prośby o zamianę dla zalogowanego użytkownika"""
    try:
        ensure_swaps_table()
        db = get_db()
        
        # Debug logging
        user_id = session.get("user_id")
        user_email = session.get("user_email", "")
        logger.info(f"api_swaps_inbox DEBUG - user_id: {user_id}, email: {user_email}")
        
        # Sprawdź czy użytkownik ma uprawnienia bossa
        user_role_row = db.execute("SELECT role FROM users WHERE id=?", (user_id,)).fetchone()
        user_role = user_role_row["role"] if user_role_row and user_role_row["role"] else "USER"
        is_boss = user_role.upper() == "ADMIN"
        logger.info(f"api_swaps_inbox DEBUG - user_role: {user_role}, is_boss: {is_boss}")
        
        if is_boss:
            # Szef widzi tylko prośby zaakceptowane przez pracowników, czekające na jego decyzję
            requests = db.execute("""
                SELECT id, from_date, from_employee, to_date, to_employee, 
                       comment_requester, recipient_status, boss_status,
                       from_shift, to_shift, is_give_request, is_ask_request, created_at,
                       COALESCE(final_status, 
                         CASE 
                           WHEN recipient_status = 'DECLINED' THEN 'ODRZUCONE'
                           WHEN recipient_status = 'ACCEPTED' AND boss_status = 'APPROVED' THEN 'ZATWIERDZONE'
                           WHEN recipient_status = 'ACCEPTED' AND boss_status = 'REJECTED' THEN 'ODRZUCONE_PRZEZ_SZEFA'
                           WHEN recipient_status = 'ACCEPTED' AND boss_status = 'PENDING' THEN 'WSTEPNIE_ZATWIERDZONE'
                           ELSE 'OCZEKUJACE'
                         END
                       ) as final_status
                FROM swap_requests 
                WHERE recipient_status = 'ACCEPTED' AND boss_status = 'PENDING'
                ORDER BY created_at DESC
            """).fetchall()
            logger.info(f"api_swaps_inbox DEBUG - found {len(requests)} boss requests (recipient_status=ACCEPTED, boss_status=PENDING)")
        else:
            # Zwykli pracownicy widzą prośby skierowane do nich lub wysłane przez nich
            me_emp = db.execute("SELECT name FROM employees WHERE user_id=?", (user_id,)).fetchone()
            my_name = me_emp["name"] if me_emp else session.get("user_name", "")
            logger.info(f"api_swaps_inbox DEBUG - my_name: {my_name}")
            
            requests = db.execute("""
                SELECT id, from_date, from_employee, to_date, to_employee, 
                       comment_requester, recipient_status, boss_status,
                       from_shift, to_shift, is_give_request, is_ask_request, created_at,
                       COALESCE(final_status, 
                         CASE 
                           WHEN recipient_status = 'DECLINED' THEN 'ODRZUCONE'
                           WHEN recipient_status = 'ACCEPTED' AND boss_status = 'APPROVED' THEN 'ZATWIERDZONE'
                           WHEN recipient_status = 'ACCEPTED' AND boss_status = 'REJECTED' THEN 'ODRZUCONE_PRZEZ_SZEFA'
                           WHEN recipient_status = 'ACCEPTED' AND boss_status = 'PENDING' THEN 'WSTEPNIE_ZATWIERDZONE'
                           ELSE 'OCZEKUJACE'
                         END
                       ) as final_status
                FROM swap_requests 
                WHERE to_employee = ? OR from_employee = ?
                ORDER BY created_at DESC
            """, (my_name, my_name)).fetchall()
            logger.info(f"api_swaps_inbox DEBUG - found {len(requests)} requests for {my_name}")
        
        response_data = {
            'items': [dict(r) for r in requests],
            'is_boss': is_boss
        }
        logger.info(f"api_swaps_inbox DEBUG - response data: {response_data}")
        
        return jsonify(response_data)
        
    except Exception as e:
        logger.error(f"Błąd podczas pobierania inbox: {e}")
        return jsonify(error="Wystąpił błąd podczas pobierania próśb"), 500

@app.post("/api/swaps/respond")
@login_required
def api_swaps_respond():
    """Odpowiada na prośbę o zamianę (ACCEPTED/DECLINED)"""
    try:
        data = safe_get_json()
        request_id = data.get("id")
        status = data.get("status")
        
        if not request_id or status not in ['ACCEPTED', 'DECLINED']:
            return jsonify(error="Nieprawidłowe dane"), 400
        
        db = get_db()
        
        # Update request status
        db.execute("""
            UPDATE swap_requests 
            SET recipient_status = ?, updated_at = datetime('now') 
            WHERE id = ?
        """, (status, request_id))
        
        # Update final status
        request_row = db.execute("SELECT recipient_status, boss_status FROM swap_requests WHERE id = ?", (request_id,)).fetchone()
        if request_row:
            final_status = calculate_final_status(status, request_row["boss_status"])
            db.execute("""
                UPDATE swap_requests 
                SET final_status = ? 
                WHERE id = ?
            """, (final_status, request_id))
        
        db.commit()
        
        logger.info(f"Użytkownik {session.get('user_email')} odpowiedział {status} na prośbę {request_id}")
        return jsonify(status="ok", message=f"Odpowiedź {status} została zapisana")
        
    except Exception as e:
        logger.error(f"Błąd podczas odpowiadania na prośbę: {e}")
        return jsonify(error="Wystąpił błąd podczas odpowiadania"), 500

def process_give_request(db, from_emp_id, to_emp_id, from_date, from_employee, to_employee):
    """Przetwarzanie prośby 'oddaj zmianę' - usuń zmianę od żądającego, dodaj do docelowego"""
    # Pobierz zmianę żądającego z danej daty
    shift_row = db.execute(
        "SELECT shift_type FROM shifts WHERE date = ? AND employee_id = ?", 
        (from_date, from_emp_id)
    ).fetchone()
    
    if shift_row:
        shift_type = shift_row["shift_type"]
        
        # Usuń zmianę od żądającego
        db.execute("DELETE FROM shifts WHERE date = ? AND employee_id = ?", (from_date, from_emp_id))
        
        # Dodaj zmianę do docelowego (usuń istniejącą jeśli jest)
        db.execute("DELETE FROM shifts WHERE date = ? AND employee_id = ?", (from_date, to_emp_id))
        db.execute("INSERT INTO shifts (date, shift_type, employee_id) VALUES (?, ?, ?)", 
                  (from_date, shift_type, to_emp_id))
        
        logger.info(f"Give request: {from_employee} oddał zmianę {shift_type} na {from_date} do {to_employee}")
    else:
        logger.warning(f"Give request: Brak zmiany dla {from_employee} na {from_date}")

def process_ask_request(db, from_emp_id, to_emp_id, to_date, from_employee, to_employee):
    """Przetwarzanie prośby 'poproś o przejęcie' - usuń zmianę od docelowego, dodaj do żądającego"""
    # Pobierz zmianę docelowego z danej daty
    shift_row = db.execute(
        "SELECT shift_type FROM shifts WHERE date = ? AND employee_id = ?", 
        (to_date, to_emp_id)
    ).fetchone()
    
    if shift_row:
        shift_type = shift_row["shift_type"]
        
        # Usuń zmianę od docelowego
        db.execute("DELETE FROM shifts WHERE date = ? AND employee_id = ?", (to_date, to_emp_id))
        
        # Dodaj zmianę do żądającego (usuń istniejącą jeśli jest)
        db.execute("DELETE FROM shifts WHERE date = ? AND employee_id = ?", (to_date, from_emp_id))
        db.execute("INSERT INTO shifts (date, shift_type, employee_id) VALUES (?, ?, ?)", 
                  (to_date, shift_type, from_emp_id))
        
        logger.info(f"Ask request: {from_employee} przejął zmianę {shift_type} na {to_date} od {to_employee}")
    else:
        logger.warning(f"Ask request: Brak zmiany dla {to_employee} na {to_date}")



@app.post("/api/swaps/boss")
@admin_required
def api_swaps_boss():
    """Zatwierdza lub odrzuca prośbę o zamianę (tylko admin)"""
    try:
        data = safe_get_json()
        request_id = data.get("id")
        status = data.get("status")
        
        if not request_id or status not in ['APPROVED', 'REJECTED']:
            return jsonify(error="Nieprawidłowe dane"), 400
        
        db = get_db()
        
        # Get request details
        request_row = db.execute("SELECT * FROM swap_requests WHERE id = ?", (request_id,)).fetchone()
        if not request_row:
            return jsonify(error="Prośba nie została znaleziona"), 404
        
        # Update boss status
        db.execute("""
            UPDATE swap_requests 
            SET boss_status = ?, updated_at = datetime('now') 
            WHERE id = ?
        """, (status, request_id))
        
        # Update final status
        final_status = calculate_final_status(request_row["recipient_status"], status)
        db.execute("""
            UPDATE swap_requests 
            SET final_status = ? 
            WHERE id = ?
        """, (final_status, request_id))
        
        # If approved, execute the swap
        if status == 'APPROVED' and request_row["recipient_status"] == 'ACCEPTED':
            from_employee = request_row["from_employee"]
            to_employee = request_row["to_employee"]
            from_date = request_row["from_date"]
            to_date = request_row["to_date"]
            is_give_request = request_row["is_give_request"]
            is_ask_request = request_row["is_ask_request"]
            
            # Get employee IDs
            from_emp_id = db.execute("SELECT id FROM employees WHERE name=?", (from_employee,)).fetchone()["id"]
            to_emp_id = db.execute("SELECT id FROM employees WHERE name=?", (to_employee,)).fetchone()["id"]
            
            if is_give_request:
                # Give request: transfer shift from requester to target
                logger.info(f"Wykonuję give_request: {from_employee} -> {to_employee} na {from_date}")
                process_give_request(db, from_emp_id, to_emp_id, from_date, from_employee, to_employee)
            elif is_ask_request:
                # Ask request: transfer shift from target to requester  
                logger.info(f"Wykonuję ask_request: {from_employee} -> {to_employee} na {to_date}")
                process_ask_request(db, from_emp_id, to_emp_id, to_date, from_employee, to_employee)
            else:
                # Regular swap: exchange shifts
                logger.info(f"Wykonuję regular_swap: {from_employee} <-> {to_employee} na daty {from_date} <-> {to_date}")
                process_regular_swap(db, from_emp_id, to_emp_id, from_date, to_date, from_employee, to_employee)
        
        db.commit()
        
        logger.info(f"Admin {session.get('user_email')} {status} prośbę {request_id}")
        return jsonify(status="ok", message=f"Prośba została {status}")
        
    except Exception as e:
        logger.error(f"Błąd podczas zatwierdzania prośby: {e}")
        return jsonify(error="Wystąpił błąd podczas zatwierdzania"), 500

@app.post("/api/swaps/clear")
@admin_required
def api_swaps_clear():
    ensure_swaps_table()
    db = get_db()
    cur = db.execute("DELETE FROM swap_requests")
    db.commit()
    return jsonify(status="ok", deleted=cur.rowcount)

@app.post("/api/notifications/test")
@login_required
def api_notifications_test():
    """Test powiadomień - endpoint do testowania"""
    try:
        return jsonify(status="ok", message="Powiadomienie testowe wysłane")
    except Exception as e:
        logger.error(f"Błąd podczas testowania powiadomień: {e}")
        return jsonify(error="Wystąpił błąd podczas testowania powiadomień"), 500

@app.get("/api/schedule/changes")
@login_required
def api_schedule_changes():
    """Pobiera ostatnie zmiany w grafiku dla powiadomień"""
    try:
        db = get_db()
        user_id = session.get("user_id")
        
        # Pobierz nazwę aktualnego użytkownika
        user_row = db.execute("SELECT name FROM users WHERE id=?", (user_id,)).fetchone()
        current_user_name = user_row["name"] if user_row else "Nieznany użytkownik"
        
        # Pobierz ostatnie zmiany z ostatnich 24 godzin
        changes = db.execute("""
            SELECT id, date, employee_name, shift_type, action, changed_by, changed_at
            FROM schedule_changes 
            WHERE changed_at > datetime('now', '-1 day')
            ORDER BY changed_at DESC
            LIMIT 50
        """).fetchall()
        
        # Konwertuj na listę słowników
        changes_list = []
        for change in changes:
            changes_list.append({
                'id': change['id'],
                'date': change['date'],
                'employee_name': change['employee_name'],
                'shift_type': change['shift_type'],
                'action': change['action'],
                'changed_by': change['changed_by'],
                'changed_at': change['changed_at']
            })
        
        return jsonify({
            'changes': changes_list,
            'current_user_name': current_user_name
        })
        
    except Exception as e:
        logger.error(f"Błąd podczas pobierania zmian w grafiku: {e}")
        return jsonify(error="Wystąpił błąd podczas pobierania zmian"), 500

# --- Unavailability API -----------------------------------------------------
@app.post("/api/unavailability")
@login_required
@rate_limit_required(max_requests=10, window=60)
def api_unavailability_create():
    """Tworzy nowe zgłoszenie niedyspozycji"""
    try:
        ensure_unavailability_table()
        data = safe_get_json()
        
        month_year = data.get("month_year", "").strip()
        selected_days = data.get("selected_days", [])
        comment = (data.get("comment") or "").strip()
        
        if not month_year or not selected_days:
            return jsonify(error="Brak miesiąca lub wybranych dni"), 400
        
        # Sprawdź czy użytkownik ma przypisanego pracownika
        db = get_db()
        me_emp = db.execute("SELECT id, name FROM employees WHERE user_id=?", (session.get("user_id"),)).fetchone()
        if not me_emp:
            return jsonify(error="Nie masz przypisanego pracownika"), 400
        
        employee_id = me_emp["id"]
        
        # Sprawdź czy nie ma już zgłoszenia na ten miesiąc
        existing = db.execute("""
            SELECT id FROM unavailability_requests 
            WHERE employee_id = ? AND month_year = ? AND status = 'PENDING'
        """, (employee_id, month_year)).fetchone()
        
        if existing:
            return jsonify(error="Masz już aktywne zgłoszenie niedyspozycji na ten miesiąc"), 400
        
        # Zapisz zgłoszenie
        import json
        selected_days_json = json.dumps(selected_days)
        
        db.execute("""
            INSERT INTO unavailability_requests(employee_id, month_year, selected_days, comment)
            VALUES (?, ?, ?, ?)
        """, (employee_id, month_year, selected_days_json, comment))
        
        db.commit()
        
        logger.info(f"Utworzono zgłoszenie niedyspozycji: {me_emp['name']} na {month_year}")
        return jsonify(status="ok", message="Zgłoszenie niedyspozycji zostało wysłane")
        
    except Exception as e:
        logger.error(f"Błąd podczas tworzenia zgłoszenia niedyspozycji: {e}")
        return jsonify(error="Wystąpił błąd podczas wysyłania zgłoszenia"), 500

@app.get("/api/unavailability/inbox")
@login_required
def api_unavailability_inbox():
    """Pobiera zgłoszenia niedyspozycji dla skrzynki"""
    try:
        ensure_unavailability_table()
        db = get_db()
        
        user_id = session.get("user_id")
        user_role_row = db.execute("SELECT role FROM users WHERE id=?", (user_id,)).fetchone()
        user_role = user_role_row["role"] if user_role_row and user_role_row["role"] else "USER"
        is_boss = user_role.upper() == "ADMIN"
        
        if is_boss:
            # Szef widzi wszystkie zgłoszenia czekające na zatwierdzenie
            requests = db.execute("""
                SELECT u.id, u.month_year, u.selected_days, u.comment, u.status, u.created_at,
                       e.name as employee_name
                FROM unavailability_requests u
                JOIN employees e ON u.employee_id = e.id
                WHERE u.status = 'PENDING'
                ORDER BY u.created_at DESC
            """).fetchall()
        else:
            # Pracownik widzi swoje zgłoszenia
            me_emp = db.execute("SELECT id FROM employees WHERE user_id=?", (user_id,)).fetchone()
            if not me_emp:
                return jsonify(items=[], is_boss=False)
            
            requests = db.execute("""
                SELECT u.id, u.month_year, u.selected_days, u.comment, u.status, u.created_at,
                       e.name as employee_name
                FROM unavailability_requests u
                JOIN employees e ON u.employee_id = e.id
                WHERE u.employee_id = ?
                ORDER BY u.created_at DESC
            """, (me_emp["id"],)).fetchall()
        
        response_data = {
            'items': [dict(r) for r in requests],
            'is_boss': is_boss
        }
        
        return jsonify(response_data)
        
    except Exception as e:
        logger.error(f"Błąd podczas pobierania zgłoszeń niedyspozycji: {e}")
        return jsonify(error="Wystąpił błąd podczas pobierania zgłoszeń"), 500

@app.post("/api/unavailability/respond")
@admin_required
def api_unavailability_respond():
    """Zatwierdza lub odrzuca zgłoszenie niedyspozycji (tylko admin)"""
    try:
        data = safe_get_json()
        request_id = data.get("id")
        status = data.get("status")
        boss_comment = (data.get("boss_comment") or "").strip()
        
        if not request_id or status not in ['APPROVED', 'REJECTED']:
            return jsonify(error="Nieprawidłowe dane"), 400
        
        db = get_db()
        
        # Pobierz dane zgłoszenia
        request_row = db.execute("""
            SELECT u.*, e.name as employee_name
            FROM unavailability_requests u
            JOIN employees e ON u.employee_id = e.id
            WHERE u.id = ?
        """, (request_id,)).fetchone()
        
        if not request_row:
            return jsonify(error="Zgłoszenie nie zostało znalezione"), 404
        
        # Aktualizuj status
        db.execute("""
            UPDATE unavailability_requests 
            SET status = ?, boss_comment = ?, updated_at = datetime('now')
            WHERE id = ?
        """, (status, boss_comment, request_id))
        
        # Jeśli zatwierdzone, dodaj wpisy "-" do tabeli shifts
        if status == 'APPROVED':
            import json
            selected_days = json.loads(request_row["selected_days"])
            employee_id = request_row["employee_id"]
            
            for date_str in selected_days:
                # Usuń istniejące wpisy dla tego pracownika w tym dniu
                db.execute("DELETE FROM shifts WHERE date = ? AND employee_id = ?", (date_str, employee_id))
                # Dodaj wpis z "-"
                db.execute("INSERT INTO shifts(date, shift_type, employee_id) VALUES (?, ?, ?)", 
                          (date_str, "-", employee_id))
        
        db.commit()
        
        logger.info(f"Admin {session.get('user_email')} {status} zgłoszenie niedyspozycji {request_id}")
        return jsonify(status="ok", message=f"Zgłoszenie zostało {status}")
        
    except Exception as e:
        logger.error(f"Błąd podczas odpowiadania na zgłoszenie niedyspozycji: {e}")
        return jsonify(error="Wystąpił błąd podczas przetwarzania zgłoszenia"), 500

# --- Employees API -------------------------------------------------------------
@app.get("/api/employees")
@admin_required
def api_employees_list():
    ensure_employees_code_column()
    db = get_db()
    rows = db.execute("SELECT id, name, code FROM employees ORDER BY name").fetchall()
    return jsonify(employees=[{"id": r["id"], "name": r["name"], "code": r["code"]} for r in rows])


@app.post("/api/employees")
@admin_required
def api_employees_add():
    ensure_employees_code_column()
    data = safe_get_json()
    code = (data.get("code") or "").strip()
    name = (data.get("name") or "").strip()
    if not code or not name:
        return jsonify(error="code and name required"), 400
    db = get_db()
    try:
        db.execute("INSERT INTO employees(name, code) VALUES (?,?)", (name, code))
        db.commit()
    except sqlite3.IntegrityError:
        return jsonify(error="employee exists or code not unique"), 409
    row = db.execute("SELECT id, name, code FROM employees WHERE code=?", (code,)).fetchone()
    return jsonify(id=row["id"], name=row["name"], code=row["code"]) , 201


@app.delete("/api/employees/<int:emp_id>")
@admin_required
def api_employees_delete(emp_id: int):
    db = get_db()
    db.execute("DELETE FROM employees WHERE id=?", (emp_id,))
    db.commit()
    return jsonify(status="ok")

@app.put("/api/employees/<int:emp_id>")
@admin_required
def api_employees_update(emp_id: int):
    ensure_employees_code_column()
    data = safe_get_json()
    code = (data.get("code") or "").strip()
    name = (data.get("name") or "").strip()
    if not code or not name:
        return jsonify(error="code and name required"), 400
    
    db = get_db()
    try:
        # Sprawdź czy pracownik istnieje
        existing = db.execute("SELECT id FROM employees WHERE id=?", (emp_id,)).fetchone()
        if not existing:
            return jsonify(error="employee not found"), 404
        
        # Sprawdź czy nowy kod nie koliduje z innymi pracownikami
        conflict = db.execute("SELECT id FROM employees WHERE code=? AND id!=?", (code, emp_id)).fetchone()
        if conflict:
            return jsonify(error="code already exists"), 409
        
        # Aktualizuj pracownika
        db.execute("UPDATE employees SET name=?, code=? WHERE id=?", (name, code, emp_id))
        db.commit()
        
        # Zwróć zaktualizowane dane
        row = db.execute("SELECT id, name, code FROM employees WHERE id=?", (emp_id,)).fetchone()
        return jsonify(id=row["id"], name=row["name"], code=row["code"])
        
    except sqlite3.IntegrityError:
        return jsonify(error="database error"), 500

# users listing for mapping
@app.get("/api/users")
@admin_required
def api_users():
    ensure_users_role_column()
    db = get_db()
    rows = db.execute("SELECT id, email, name, role FROM users ORDER BY email").fetchall()
    return jsonify(users=[{"id":r["id"],"email":r["email"],"name":r["name"],"role":r["role"]} for r in rows])

# link employee to user
@app.post("/api/employees/link")
@admin_required
def api_employees_link():
    """Linkuje pracownika z użytkownikiem"""
    data = safe_get_json()
    employee_id = data.get("employee_id")
    user_id = data.get("user_id")
    
    if not employee_id or not user_id:
        return jsonify(error="employee_id and user_id required"), 400
    
    try:
        db = get_db()
        db.execute("UPDATE employees SET user_id=? WHERE id=?", (user_id, employee_id))
        db.commit()
        return jsonify(status="ok")
    except sqlite3.Error as e:
        logger.error(f"Błąd linkowania pracownika: {e}")
        return jsonify(error="database error"), 500

# Web Push Notifications API
@app.post("/api/push/subscribe")
@login_required
def subscribe_push():
    """Zapisuje subskrypcję push dla użytkownika"""
    try:
        subscription = request.get_json()
        user_id = session.get("user_id")
        
        logger.info(f"Push subscribe - user_id: {user_id}, subscription: {subscription}")
        
        if not subscription:
            logger.error("Brak danych subskrypcji")
            return jsonify(error="Brak danych subskrypcji"), 400
        
        if not user_id:
            logger.error("Brak user_id w sesji")
            return jsonify(error="Brak user_id w sesji"), 400
        
        # Walidacja danych subskrypcji
        required_fields = ['endpoint', 'keys']
        if not all(field in subscription for field in required_fields):
            return jsonify(error="Nieprawidłowe dane subskrypcji"), 400
        
        if 'keys' not in subscription or not subscription['keys']:
            return jsonify(error="Brak kluczy w subskrypcji"), 400
        
        if save_push_subscription(user_id, subscription):
            return jsonify(message="Subskrypcja zapisana pomyślnie")
        else:
            return jsonify(error="Błąd zapisywania subskrypcji"), 500
            
    except Exception as e:
        logger.error(f"Błąd subskrypcji push: {e}")
        return jsonify(error="Błąd serwera"), 500

@app.get("/api/push/vapid-key")
def get_vapid_key():
    """Zwraca klucz publiczny VAPID dla frontendu"""
    try:
        public_key = app.config.get('VAPID_PUBLIC_KEY')
        if not public_key:
            return jsonify(error="Brak konfiguracji VAPID"), 500
        
        return jsonify(public_key=public_key)
    except Exception as e:
        logger.error(f"Błąd pobierania klucza VAPID: {e}")
        return jsonify(error="Błąd serwera"), 500

@app.post("/api/push/send")
@login_required
@rate_limit_required(max_requests=5, window=60)
def send_push_notification_api():
    """Wysyła powiadomienie push do użytkownika (test)"""
    try:
        data = request.get_json()
        title = data.get('title', 'Test powiadomienia')
        body = data.get('body', 'To jest testowe powiadomienie')
        
        user_id = session.get("user_id")
        subscription = get_push_subscriptions(user_id)
        
        if not subscription:
            return jsonify(error="Brak subskrypcji push"), 400
        
        if send_push_notification(subscription, title, body):
            return jsonify(message="Powiadomienie wysłane pomyślnie")
        else:
            return jsonify(error="Błąd wysyłania powiadomienia"), 500
            
    except Exception as e:
        logger.error(f"Błąd wysyłania powiadomienia push: {e}")
        return jsonify(error="Błąd serwera"), 500

@app.post("/api/push/send-to-all")
@login_required
def send_push_to_all():
    """Wysyła powiadomienie push do wszystkich użytkowników (tylko admin)"""
    try:
        # Sprawdź czy użytkownik jest administratorem
        user_email = session.get("user_email")
        if not is_admin_user(user_email):
            return jsonify(error="Brak uprawnień"), 403
        
        data = request.get_json()
        title = data.get('title', 'Ogłoszenie')
        body = data.get('body', 'Nowe ogłoszenie w systemie')
        
        subscriptions = get_push_subscriptions()
        success_count = 0
        
        for user_id, subscription in subscriptions:
            if send_push_notification(subscription, title, body):
                success_count += 1
        
        return jsonify(
            message=f"Powiadomienia wysłane do {success_count} użytkowników",
            sent_count=success_count,
            total_count=len(subscriptions)
        )
        
    except Exception as e:
        logger.error(f"Błąd wysyłania powiadomień do wszystkich: {e}")
        return jsonify(error="Błąd serwera"), 500

@app.post("/api/push/cleanup")
@login_required
def cleanup_push_subscriptions():
    """Czyści wygasłe subskrypcje push (tylko admin)"""
    try:
        # Sprawdź czy użytkownik jest administratorem
        user_email = session.get("user_email")
        if not is_admin_user(user_email):
            return jsonify(error="Brak uprawnień"), 403
        
        cleanup_expired_subscriptions()
        return jsonify(message="Czyszczenie wygasłych subskrypcji zakończone")
        
    except Exception as e:
        logger.error(f"Błąd czyszczenia subskrypcji: {e}")
        return jsonify(error="Błąd serwera"), 500

# Authentication routes
@app.get("/signin")
def signin():
    """Strona logowania"""
    return render_template("signin.html")

@app.get("/offline")
def offline():
    """Strona offline dla PWA"""
    return render_template("offline.html")

@app.get("/static/manifest.json")
def manifest():
    """Manifest PWA"""
    return send_from_directory('static', 'manifest.json', mimetype='application/manifest+json')

@app.get("/static/sw.js")
def service_worker():
    """Service Worker dla PWA"""
    return send_from_directory('static', 'sw.js', mimetype='application/javascript')

@app.get("/favicon.ico")
def favicon():
    """Favicon dla przeglądarki"""
    return send_from_directory('static', 'favicon.ico', mimetype='image/x-icon')

@app.get("/login")
def login():
    """Przekierowanie do Google OAuth"""
    # W produkcji zawsze używaj HTTPS
    if app.config.get('PREFERRED_URL_SCHEME') == 'https':
        # Wymuś HTTPS dla callback URL
        redirect_uri = url_for('auth_callback', _external=True, _scheme='https')
        app.logger.info(f"🔒 HTTPS: Generated redirect_uri = {redirect_uri}")
    else:
        redirect_uri = url_for('auth_callback', _external=True)
        app.logger.info(f"🔓 HTTP: Generated redirect_uri = {redirect_uri}")
    
    return google.authorize_redirect(redirect_uri)

@app.get("/auth/callback")
@app.get("/authorize")  # Alternative route for compatibility
def auth_callback():
    """Callback po autoryzacji Google"""
    # Sprawdź czy używamy HTTPS w produkcji
    if app.config.get('PREFERRED_URL_SCHEME') == 'https':
        is_https = (
            request.headers.get('X-Forwarded-Proto') == 'https' or
            request.headers.get('X-Forwarded-Scheme') == 'https' or
            request.is_secure
        )
        if not is_https:
            app.logger.warning(f"⚠️ Próba dostępu do callback przez HTTP w produkcji")
            # Przekieruj na HTTPS
            url = request.url.replace('http://', 'https://', 1)
            return redirect(url, code=301)
    
    try:
        token = google.authorize_access_token()
        user_info = token.get('userinfo')
        
        if not user_info:
            logger.error("Nie otrzymano informacji o użytkowniku z Google")
            return redirect(url_for('signin'))
        
        email = user_info.get('email', '').lower()
        name = user_info.get('name', '')
        picture = user_info.get('picture', '')
        google_sub = user_info.get('sub', '')
        
        logger.info(f"Próba logowania użytkownika: {email}")
        
        # Sprawdź whitelist
        whitelist = load_whitelist()
        if whitelist and email not in whitelist:
            logger.warning(f"Użytkownik {email} nie jest na liście dozwolonych")
            return render_template("signin.html", error="Twój email nie jest autoryzowany do korzystania z tej aplikacji")
        
        # Zapisz lub zaktualizuj użytkownika w bazie
        ensure_users_role_column()
        db = get_db()
        
        # Sprawdź czy użytkownik już istnieje
        user = db.execute("SELECT id FROM users WHERE email=?", (email,)).fetchone()
        
        if user:
            # Aktualizuj istniejącego użytkownika
            db.execute("""
                UPDATE users SET name=?, picture=?, google_sub=? WHERE email=?
            """, (name, picture, google_sub, email))
            user_id = user["id"]
        else:
            # Utwórz nowego użytkownika
            cursor = db.execute("""
                INSERT INTO users (email, name, picture, google_sub, role) 
                VALUES (?, ?, ?, ?, 'USER')
            """, (email, name, picture, google_sub))
            user_id = cursor.lastrowid
        
        db.commit()
        
        # Ustaw sesję
        session.permanent = True
        session["user_id"] = user_id
        session["user_email"] = email
        session["user_name"] = name
        
        logger.info(f"Użytkownik {email} zalogowany pomyślnie")
        return redirect(url_for('index'))
        
    except Exception as e:
        logger.error(f"Błąd podczas autoryzacji: {e}")
        return render_template("signin.html", error="Wystąpił błąd podczas logowania. Spróbuj ponownie.")

@app.post("/logout")
def logout():
    """Wylogowanie użytkownika"""
    user_email = session.get("user_email", "nieznany")
    session.clear()
    logger.info(f"Użytkownik {user_email} wylogowany")
    return redirect(url_for('signin'))

# Funkcje pomocnicze dla strony głównej
def get_calendar_navigation(year, month):
    """
    Oblicza daty nawigacji kalendarza (poprzedni/następny miesiąc).
    """
    if month == 1:
        prev_year, prev_month = year - 1, 12
    else:
        prev_year, prev_month = year, month - 1
        
    if month == 12:
        next_year, next_month = year + 1, 1
    else:
        next_year, next_month = year, month + 1
    
    return (prev_year, prev_month), (next_year, next_month)

def get_month_label(year, month):
    """
    Tworzy etykietę miesiąca w języku polskim.
    """
    month_names = ['', 'Styczeń', 'Luty', 'Marzec', 'Kwiecień', 'Maj', 'Czerwiec', 
                  'Lipiec', 'Sierpień', 'Wrzesień', 'Październik', 'Listopad', 'Grudzień']
    return f"{month_names[month]} {year}"

def get_today_shifts(db):
    """
    Pobiera dzisiejsze zmiany z bazy danych i grupuje je według typów.
    """
    today = dt.datetime.now().strftime('%Y-%m-%d')
    
    # Pobierz dzisiejsze zmiany
    today_shifts = db.execute("""
        SELECT e.name, s.shift_type
        FROM shifts s 
        JOIN employees e ON s.employee_id = e.id 
        WHERE s.date = ?
    """, (today,)).fetchall()
    
    # Przygotuj shifts_today dynamicznie dla wszystkich typów zmian
    shifts_today = {}
    for shift in today_shifts:
        shift_type = shift["shift_type"]
        
        # Mapuj międzyzmiany (P 10-22) do kategorii POPOLUDNIOWKA
        if shift_type and shift_type.startswith('P '):
            category = "POPOLUDNIOWKA"
            # Wyciągnij godziny z międzyzmiany (P 10-22 -> 10-22)
            hours = shift_type[2:] if len(shift_type) > 2 else ""
            display_name = f"{shift['name']} ({hours})" if hours else shift['name']
        else:
            category = shift_type
            display_name = shift["name"]
            
        if category not in shifts_today:
            shifts_today[category] = []
        shifts_today[category].append(display_name)
    
    # Upewnij się, że standardowe typy istnieją (dla kompatybilności z frontendem)
    if "DNIOWKA" not in shifts_today:
        shifts_today["DNIOWKA"] = []
    if "NOCKA" not in shifts_today:
        shifts_today["NOCKA"] = []
    if "POPOLUDNIOWKA" not in shifts_today:
        shifts_today["POPOLUDNIOWKA"] = []
    
    # Dodaj małe litery dla template (kompatybilność z frontendem)
    shifts_today["dniowka"] = shifts_today.get("DNIOWKA", [])
    shifts_today["nocka"] = shifts_today.get("NOCKA", [])
    shifts_today["popoludniowka"] = shifts_today.get("POPOLUDNIOWKA", [])
    
    return shifts_today

# Main page route
@app.get("/")
@login_required
def index():
    """Główna strona aplikacji - wymagane logowanie"""
    try:
        # Zapewnij że wszystkie tabele istnieją
        ensure_unavailability_table()
        
        # Pobierz parametry kalendarza
        year = int(request.args.get('year', dt.datetime.now().year))
        month = int(request.args.get('month', dt.datetime.now().month))
        
        # Oblicz nawigację kalendarza
        (prev_year, prev_month), (next_year, next_month) = get_calendar_navigation(year, month)
        
        # Utwórz etykietę miesiąca
        month_label = get_month_label(year, month)
        
        # Pobierz dane z bazy danych
        db = get_db()
        employees = db.execute("SELECT id, name FROM employees ORDER BY name").fetchall()
        shifts_today = get_today_shifts(db)
        
        # Generate calendar days for the month with proper structure
        import datetime as dt_module
        
        # Polish holidays (basic ones - you can expand this)
        polish_holidays = {
            (1, 1): "Nowy Rok",
            (1, 6): "Święto Trzech Króli", 
            (5, 1): "Święto Pracy",
            (5, 3): "Święto Konstytucji",
            (8, 15): "Wniebowzięcie NMP",
            (11, 1): "Wszystkich Świętych",
            (11, 11): "Święto Niepodległości",
            (12, 25): "Boże Narodzenie",
            (12, 26): "Św. Szczepana"
        }
        
        # Day names in Polish
        day_names = ['Pon', 'Wt', 'Śr', 'Czw', 'Pt', 'Sob', 'Nie']
        
        calendar_days = []
        month_calendar = calendar.monthcalendar(year, month)
        
        for week in month_calendar:
            calendar_week = []
            for day in week:
                if day == 0:
                    calendar_week.append(None)
                else:
                    date_str = f"{year:04d}-{month:02d}-{day:02d}"
                    date_obj = dt_module.date(year, month, day)
                    
                    # Check if it's weekend or holiday
                    is_weekend = date_obj.weekday() >= 5  # Saturday=5, Sunday=6
                    is_holiday = (month, day) in polish_holidays
                    is_off = is_weekend or is_holiday
                    
                    # Get shifts for this day
                    day_shifts = db.execute("""
                        SELECT e.name, s.shift_type
                        FROM shifts s 
                        JOIN employees e ON s.employee_id = e.id 
                        WHERE s.date = ?
                    """, (date_str,)).fetchall()
                    
                    # Inicjalizuj słownik zmian dynamicznie
                    shifts_dict = {}
                    for shift in day_shifts:
                        shift_type = shift["shift_type"]
                        if shift_type not in shifts_dict:
                            shifts_dict[shift_type] = []
                        shifts_dict[shift_type].append(shift["name"])
                    
                    # Upewnij się, że standardowe typy istnieją (dla kompatybilności z frontendem)
                    if "DNIOWKA" not in shifts_dict:
                        shifts_dict["DNIOWKA"] = []
                    if "NOCKA" not in shifts_dict:
                        shifts_dict["NOCKA"] = []
                    if "POPOLUDNIOWKA" not in shifts_dict:
                        shifts_dict["POPOLUDNIOWKA"] = []
                    
                    day_data = {
                        "day": day,
                        "dd": f"{day:02d}",  # Format as 01, 02, etc.
                        "date": date_str,
                        "iso": date_str,
                        "abbr": day_names[date_obj.weekday()],  # Mon=0, so day_names[0]='Pon'
                        "is_off": is_off,
                        "is_weekend": is_weekend,
                        "is_holiday": is_holiday,
                        "holiday_name": polish_holidays.get((month, day), ""),
                        "shifts": shifts_dict
                    }
                    
                    calendar_week.append(day_data)
            calendar_days.append(calendar_week)

        # Get current employee name for this user
        me_emp = db.execute("SELECT name FROM employees WHERE user_id=?", (session.get("user_id"),)).fetchone()
        current_emp_name = me_emp["name"] if me_emp else session.get("user_name", "")
        
        # Check if user is admin
        user_role_row = db.execute("SELECT is_special FROM employees WHERE user_id=?", (session.get("user_id"),)).fetchone()
        is_admin = user_role_row and user_role_row["is_special"] == 1
        
        # Debug: sprawdź dane przed renderowaniem
        logger.info(f"Renderowanie strony - użytkownik: {session.get('user_name')}")
        logger.info(f"Liczba pracowników: {len(employees)}")
        logger.info(f"Liczba tygodni kalendarza: {len(calendar_days)}")
        
        # Konwertuj sqlite Row objects na dict żeby uniknąć problemów z serializacją
        # i dodaj licznik zmian dla każdego pracownika
        employees_clean = []
        for e in employees:
            # Policz zmiany dla tego pracownika w bieżącym miesiącu
            shift_count = db.execute("""
                SELECT COUNT(*) as count 
                FROM shifts 
                WHERE employee_id = ? 
                AND date LIKE ?
            """, (e["id"], f"{year:04d}-{month:02d}-%")).fetchone()
            
            count = shift_count["count"] if shift_count else 0
            display_name = f"{e['name']} ({count})" if e["name"] else ""
            
            employees_clean.append({
                "id": int(e["id"]) if e["id"] is not None else 0,
                "name": str(e["name"]) if e["name"] is not None else "",
                "display_name": display_name,
                "shift_count": count
            })
        
        # Clean calendar_days and ensure proper types
        calendar_days_clean = []
        for week in calendar_days:
            week_clean = []
            for day in week:
                if day is None:
                    week_clean.append(None)
                else:
                    day_clean = {
                        "day": int(day["day"]) if day["day"] is not None else 0,
                        "dd": str(day["dd"]) if day["dd"] is not None else "",
                        "date": str(day["date"]) if day["date"] is not None else "",
                        "iso": str(day["iso"]) if day["iso"] is not None else "",
                        "abbr": str(day["abbr"]) if day["abbr"] is not None else "",
                        "is_off": bool(day["is_off"]) if day["is_off"] is not None else False,
                        "is_weekend": bool(day["is_weekend"]) if day["is_weekend"] is not None else False,
                        "is_holiday": bool(day["is_holiday"]) if day["is_holiday"] is not None else False,
                        "holiday_name": str(day["holiday_name"]) if day["holiday_name"] is not None else "",
                        "shifts": {
                            shift_type: [str(name) for name in employees_list] 
                            for shift_type, employees_list in day["shifts"].items()
                        }
                    }
                    week_clean.append(day_clean)
            calendar_days_clean.append(week_clean)
        
        # Przygotuj schedule_rows - płaska lista dni z danymi
        schedule_rows = []
        for week in calendar_days_clean:
            for day in week:
                if day is not None:
                    schedule_rows.append(day)
        
        # Przygotuj shifts_by_date dla JavaScript
        shifts_by_date = {}
        for week in calendar_days_clean:
            for day in week:
                if day is not None:
                    date_str = day["date"]
                    shifts_by_date[date_str] = {}
                    for emp in employees_clean:
                        emp_name = emp["name"]
                        # Sprawdź czy pracownik ma zmianę tego dnia (dowolny typ)
                        for shift_type, employees_list in day["shifts"].items():
                            if emp_name in employees_list:
                                shifts_by_date[date_str][emp_name] = shift_type
                                break  # Pracownik może mieć tylko jedną zmianę dziennie
        
        response = make_response(render_template("index.html", 
            shifts_today=shifts_today,
            month_label=str(month_label),
            prev_year=int(prev_year), prev_month=int(prev_month),
            next_year=int(next_year), next_month=int(next_month),
            view_year=int(year), view_month=int(month),
            employees=employees_clean,
            calendar_days=calendar_days_clean,
            schedule_rows=schedule_rows,
            current_user=str(session.get("user_name", "Użytkownik")),
            current_emp_name=str(current_emp_name) if current_emp_name else "",
            shifts_by_date=shifts_by_date,
            is_admin=is_admin
        ))
        
        response.headers['Cache-Control'] = 'no-cache, no-store, must-revalidate, max-age=0'
        response.headers['Pragma'] = 'no-cache'
        response.headers['Expires'] = '0'
        response.headers['Last-Modified'] = str(dt.datetime.now())
        response.headers['ETag'] = f'"{hash(str(dt.datetime.now()))}"'
        logger.info(f"Główna strona załadowana dla użytkownika {session.get("user_email")}")
        return response

    except Exception as e:
        logger.error(f"Błąd podczas ładowania głównej strony: {e}")
        abort(500, "Błąd podczas ładowania strony")

# --- Health check i debug -----------------------------------------------------
@app.get("/healthz")
def healthz():
    """Health check endpoint"""
    return jsonify(status="ok", timestamp=dt.datetime.now().isoformat())

@app.get("/debug/env")
def debug_env():
    """Debug endpoint - sprawdza konfigurację środowiska"""
    return jsonify(
        has_client_id=bool(os.environ.get("GOOGLE_CLIENT_ID")),
        has_client_secret=bool(os.environ.get("GOOGLE_CLIENT_SECRET")),
        has_secret_key=bool(os.environ.get("SECRET_KEY")),
        environment=os.environ.get("FLASK_ENV", "development")
    )

# --- Excel Export API -------------------------------------------------------------
@app.get("/api/export/excel")
def api_export_excel():
    """Eksportuje grafik do pliku Excel - imiona u góry, dni po lewej"""
    try:
        logger.info(f"Excel export - user_id: {session.get('user_id')}, user_email: {session.get('user_email')}")
        logger.info(f"Excel export - session data: {dict(session)}")
        
        # Sprawdź czy użytkownik jest zalogowany
        if not session.get("user_id"):
            logger.warning("Excel export - użytkownik nie jest zalogowany")
            return redirect(url_for("signin"))
        
        # Sprawdź uprawnienia admin
        db = get_db()
        user_role_row = db.execute("SELECT role FROM users WHERE id=?", (session["user_id"],)).fetchone()
        user_role = user_role_row["role"] if user_role_row and user_role_row["role"] else "USER"
        
        if user_role != "ADMIN":
            logger.warning(f"Excel export - użytkownik {session['user_id']} nie ma uprawnień admin (rola: {user_role})")
            return jsonify(error="Brak uprawnień administratora"), 403
        
        logger.info(f"Excel export - użytkownik {session['user_id']} ma uprawnienia admin")
        
        # Import bibliotek
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
        from datetime import datetime
        import io
        import calendar
        
        # Pobierz parametry miesiąca z URL
        year = int(request.args.get('year', datetime.now().year))
        month = int(request.args.get('month', datetime.now().month))
        
        db = get_db()
        
        # Pobierz listę pracowników z licznikami zmian
        employees = db.execute("SELECT id, name FROM employees ORDER BY name").fetchall()
        
        # Dodaj liczniki zmian dla każdego pracownika
        employees_with_counts = []
        for employee in employees:
            # Policz zmiany dla tego pracownika w bieżącym miesiącu
            shift_count = db.execute("""
                SELECT COUNT(*) as count 
                FROM shifts 
                WHERE employee_id = ? 
                AND date LIKE ?
            """, (employee["id"], f"{year:04d}-{month:02d}-%")).fetchone()
            
            count = shift_count["count"] if shift_count else 0
            display_name = f"{employee['name']} ({count})" if employee["name"] else ""
            
            employees_with_counts.append({
                "id": employee["id"],
                "name": employee["name"],
                "display_name": display_name,
                "shift_count": count
            })
        
        # Pobierz wszystkie zmiany dla danego miesiąca
        month_pattern = f"{year:04d}-{month:02d}-%"
        shifts_data = db.execute("""
            SELECT s.date, s.shift_type, e.name as employee_name, e.id as employee_id
            FROM shifts s
            JOIN employees e ON s.employee_id = e.id
            WHERE s.date LIKE ?
            ORDER BY s.date, e.name
        """, (month_pattern,)).fetchall()
        
        # Utwórz słownik zmian dla łatwego dostępu
        shifts_dict = {}
        for shift in shifts_data:
            date = shift['date']
            employee_name = shift['employee_name']
            if date not in shifts_dict:
                shifts_dict[date] = {}
            shifts_dict[date][employee_name] = shift['shift_type']
        
        # Utwórz nowy skoroszyt
        wb = Workbook()
        ws = wb.active
        ws.title = "Grafik SP4600"
        
        # Style
        header_font = Font(bold=True, color="FFFFFF", size=12)
        header_fill = PatternFill(start_color="2E7D32", end_color="2E7D32", fill_type="solid")
        date_font = Font(bold=True, size=11)
        date_fill = PatternFill(start_color="E8F5E8", end_color="E8F5E8", fill_type="solid")
        border = Border(
            left=Side(style='thin', color='000000'),
            right=Side(style='thin', color='000000'),
            top=Side(style='thin', color='000000'),
            bottom=Side(style='thin', color='000000')
        )
        center_alignment = Alignment(horizontal='center', vertical='center')
        
        # Nazwy miesięcy po polsku
        month_names = ['', 'Styczeń', 'Luty', 'Marzec', 'Kwiecień', 'Maj', 'Czerwiec', 
                      'Lipiec', 'Sierpień', 'Wrzesień', 'Październik', 'Listopad', 'Grudzień']
        month_label = f"{month_names[month]} {year}"
        
        # Tytuł główny
        ws['A1'] = f"Grafik zmian pracowników SP4600 - {month_label}"
        ws['A1'].font = Font(bold=True, size=16)
        
        # Scal komórki dla tytułu
        total_cols = len(employees_with_counts) + 3  # +3 dla DZIEŃ, DATA, LICZNIK
        ws.merge_cells(f'A1:{get_column_letter(total_cols)}1')
        
        # Nagłówek pierwszej kolumny (dni)
        ws['A3'] = "DZIEŃ"
        ws['A3'].font = header_font
        ws['A3'].fill = header_fill
        ws['A3'].alignment = center_alignment
        ws['A3'].border = border
        
        # Nagłówek drugiej kolumny (data)
        ws['B3'] = "DATA"
        ws['B3'].font = header_font
        ws['B3'].fill = header_fill
        ws['B3'].alignment = center_alignment
        ws['B3'].border = border
        
        # Nagłówek trzeciej kolumny (licznik - tylko dla admin)
        ws['C3'] = "LICZNIK"
        ws['C3'].font = header_font
        ws['C3'].fill = header_fill
        ws['C3'].alignment = center_alignment
        ws['C3'].border = border
        
        # Nagłówki kolumn - imiona pracowników z licznikami
        col = 4  # Zaczynamy od kolumny D (4)
        for employee in employees_with_counts:
            cell = ws.cell(row=3, column=col, value=employee['display_name'])
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_alignment
            cell.border = border
            col += 1
        
        # Generuj wszystkie dni miesiąca
        days_in_month = calendar.monthrange(year, month)[1]
        day_names = ['Pon', 'Wt', 'Śr', 'Czw', 'Pt', 'Sob', 'Nie']
        
        # Wypełnij dane - każdy wiersz to jeden dzień
        row = 4
        for day in range(1, days_in_month + 1):
            date_str = f"{year:04d}-{month:02d}-{day:02d}"
            date_obj = datetime.strptime(date_str, '%Y-%m-%d')
            day_name = day_names[date_obj.weekday()]
            
            # Pierwsza kolumna - dzień miesiąca
            cell = ws.cell(row=row, column=1, value=f"{day:02d} ({day_name})")
            cell.font = date_font
            cell.fill = date_fill
            cell.alignment = center_alignment
            cell.border = border
            
            # Druga kolumna - data
            cell = ws.cell(row=row, column=2, value=date_str)
            cell.font = date_font
            cell.fill = date_fill
            cell.alignment = center_alignment
            cell.border = border
            
            # Trzecia kolumna - licznik zmian (tylko dla admin)
            # Policz zmiany dla tego dnia
            day_shifts = shifts_dict.get(date_str, {})
            dniowka_count = sum(1 for shift in day_shifts.values() if 'DNIOWKA' in shift)
            popoludniowka_count = sum(1 for shift in day_shifts.values() if 'POPOLUDNIOWKA' in shift or shift.startswith('P '))
            nocka_count = sum(1 for shift in day_shifts.values() if 'NOCKA' in shift)
            
            licznik_text = f"D:{dniowka_count} P:{popoludniowka_count} N:{nocka_count}"
            cell = ws.cell(row=row, column=3, value=licznik_text)
            cell.font = date_font
            cell.fill = date_fill
            cell.alignment = center_alignment
            cell.border = border
            
            # Wypełnij zmiany dla każdego pracownika w tym dniu
            col = 4  # Zaczynamy od kolumny D (4)
            for employee in employees_with_counts:
                employee_name = employee['name']
                
                # Sprawdź czy pracownik ma zmianę w tym dniu
                shift_type = shifts_dict.get(date_str, {}).get(employee_name, "")
                
                # Kolorowanie w zależności od typu zmiany
                if shift_type:
                    if 'DNIOWKA' in shift_type:
                        fill_color = "E3F2FD"  # Jasny niebieski
                    elif 'NOCKA' in shift_type:
                        fill_color = "F3E5F5"  # Jasny fioletowy
                    elif 'POPOLUDNIOWKA' in shift_type or shift_type.startswith('P '):
                        fill_color = "E8F5E8"  # Jasny zielony
                    else:
                        fill_color = "FFF3E0"  # Jasny pomarańczowy
                else:
                    fill_color = "FFFFFF"  # Biały
                
                cell = ws.cell(row=row, column=col, value=shift_type)
                cell.fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
                cell.border = border
                cell.alignment = center_alignment
                col += 1
            
            row += 1
        
        # Dostosuj szerokość kolumn
        ws.column_dimensions['A'].width = 18  # Kolumna z datami
        ws.column_dimensions['B'].width = 12  # Kolumna z datą
        ws.column_dimensions['C'].width = 15  # Kolumna z licznikiem
        for col in range(4, len(employees_with_counts) + 4):
            ws.column_dimensions[get_column_letter(col)].width = 15
        
        # Zamroź pierwsze wiersze (tytuł i nagłówki)
        ws.freeze_panes = 'D4'  # Zamroź kolumny A, B, C
        
        # Zapisz do BytesIO
        excel_buffer = io.BytesIO()
        wb.save(excel_buffer)
        excel_buffer.seek(0)
        
        # Przygotuj nazwę pliku (bez polskich znaków)
        safe_month_label = month_label.replace('ą', 'a').replace('ć', 'c').replace('ę', 'e').replace('ł', 'l').replace('ń', 'n').replace('ó', 'o').replace('ś', 's').replace('ź', 'z').replace('ż', 'z')
        filename = f"grafik_sp4600_{safe_month_label.replace(' ', '_')}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        
        logger.info(f"Excel export completed: {filename}")
        
        response = make_response(excel_buffer.getvalue())
        response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        response.headers['Content-Disposition'] = f'attachment; filename*=UTF-8\'\'{filename}'
        response.headers['Cache-Control'] = 'no-cache, no-store, must-revalidate'
        response.headers['Pragma'] = 'no-cache'
        response.headers['Expires'] = '0'
        
        return response
        
    except ImportError:
        logger.error("Biblioteka openpyxl nie jest zainstalowana")
        return jsonify(error="Biblioteka openpyxl nie jest zainstalowana. Zainstaluj: pip install openpyxl"), 500
    except Exception as e:
        logger.error(f"Błąd podczas eksportu do Excel: {e}")
        return jsonify(error=f"Wystąpił błąd podczas eksportu do Excel: {str(e)}"), 500

def send_schedule_change_notifications(changes, changed_by_email):
    """
    Wysyła powiadomienia push o zmianach w grafiku
    """
    if not WEB_PUSH_AVAILABLE:
        logger.warning("Web Push nie jest dostępne - pomijam wysyłanie powiadomień")
        return
    
    try:
        logger.info(f"Wysyłanie powiadomień o {len(changes)} zmianach przez {changed_by_email}")
        
        # Grupuj zmiany według pracownika
        changes_by_employee = {}
        for change in changes:
            employee_name = change.get('employee') or change.get('name')  # Support both formats
            logger.info(f"Przetwarzanie zmiany dla pracownika: {employee_name} (zmienione przez: {changed_by_email})")
            if employee_name and employee_name != changed_by_email:
                if employee_name not in changes_by_employee:
                    changes_by_employee[employee_name] = []
                changes_by_employee[employee_name].append(change)
                logger.info(f"Dodano zmianę dla {employee_name}")
            else:
                logger.info(f"Pominięto zmianę dla {employee_name} (to sam użytkownik lub brak nazwy)")
        
        logger.info(f"Znaleziono {len(changes_by_employee)} pracowników do powiadomienia: {list(changes_by_employee.keys())}")
        
        # Wyślij powiadomienia do każdego pracownika
        for employee_name, employee_changes in changes_by_employee.items():
            # Znajdź user_id pracownika
            db = get_db()
            cursor = db.cursor()
            cursor.execute("SELECT id FROM users WHERE name = ?", (employee_name,))
            result = cursor.fetchone()
            
            if result:
                user_id = result[0]
                subscriptions = get_push_subscriptions(user_id)
                
                if subscriptions:
                    # Przygotuj wiadomość
                    change_count = len(employee_changes)
                    if change_count == 1:
                        change = employee_changes[0]
                        title = "Zmiana w grafiku"
                        body = f"Zmieniono Twój grafik na {change['date']} - {change['shift_type']}"
                    else:
                        title = "Zmiany w grafiku"
                        body = f"Masz {change_count} nowych zmian w grafiku"
                    
                    # Wyślij powiadomienie
                    for subscription in subscriptions:
                        send_push_notification(subscription, title, body, {
                            'type': 'schedule_change',
                            'employee': employee_name,
                            'changes_count': change_count
                        })
                    
                    logger.info(f"Wysłano powiadomienie o zmianach w grafiku dla {employee_name}")
                else:
                    logger.info(f"Brak subskrypcji push dla {employee_name}")
            else:
                logger.warning(f"Nie znaleziono użytkownika: {employee_name}")
                
    except Exception as e:
        logger.error(f"Błąd wysyłania powiadomień o zmianach w grafiku: {e}")

# --- Main --------------------------------------------------------------------
if __name__ == "__main__":
    logger.info("Uruchamianie aplikacji Flask...")
    
    # Zainicjalizuj wszystkie tabele przy starcie aplikacji
    with app.app_context():
        try:
            ensure_users_role_column()
            ensure_employees_user_id_column()
            ensure_employees_code_column()
            ensure_swaps_table()
            ensure_unavailability_table()
            ensure_schedule_changes_table()
            ensure_push_subscriptions_table()
            logger.info("Wszystkie tabele bazy danych zostały zainicjalizowane")
        except Exception as e:
            logger.error(f"Błąd podczas inicjalizacji tabel: {e}")
    
    app.run(debug=False, host='localhost', port=5000)
